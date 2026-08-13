from __future__ import annotations
import json
from datetime import date, datetime
from .. import client
from apps.historial import repo as historial_repo


class PeriodoFuturoEncolado(Exception):
    """entrada_documento levanta esto cuando el documento pertenece a un
    periodo POSTERIOR al periodo de proceso abierto de CxP: en vez de
    insertarlo (lo cual lo "adelantaria" a un mes que aun no se abre) se
    guarda en TCXP_DOCUMENTO_COLA y se materializa solo cuando el cierre
    avance mes_proceso hasta ese mes. Los callers (vista manual, espejo INV)
    lo tratan como exito-diferido, NO como error."""
    def __init__(self, cola_id: int, periodo_objetivo: str, periodo_proceso: str):
        self.cola_id = cola_id
        self.periodo_objetivo = periodo_objetivo
        self.periodo_proceso = periodo_proceso
        super().__init__(
            f"Documento del periodo {periodo_objetivo} encolado; el periodo "
            f"de proceso CxP en curso es {periodo_proceso}. Se materializara "
            f"automaticamente al abrir {periodo_objetivo}.")


def _periodo_estado(no_cia: str, punto: str, fecha_yyyy_mm_dd: str):
    """Devuelve (estado, periodo_proceso_str) donde estado es
    'ABIERTO' | 'FUTURO' | 'CERRADO' comparando el periodo de la fecha del
    documento contra TCXP_PUNTO.ano_proceso/mes_proceso. Si el punto no
    tiene fila de proceso, devuelve ('ABIERTO', None) para no bloquear."""
    periodo_docu = (fecha_yyyy_mm_dd or '')[:7]  # 'YYYY-MM'
    row = client.fetch_one(
        "SELECT NVL(ano_proceso,0), NVL(mes_proceso,0) FROM CXP.TCXP_PUNTO "
        "WHERE no_cia=:1 AND punto=:2", [no_cia, punto])
    if not row or not row[0] or not row[1] or len(periodo_docu) != 7:
        return 'ABIERTO', None
    periodo_proceso = '{:04d}-{:02d}'.format(int(row[0]), int(row[1]))
    if periodo_docu == periodo_proceso:
        return 'ABIERTO', periodo_proceso
    if periodo_docu > periodo_proceso:
        return 'FUTURO', periodo_proceso
    return 'CERRADO', periodo_proceso


def _encolar_documento(d: dict, origen: str = 'MANUAL') -> int:
    """Guarda el payload del documento en TCXP_DOCUMENTO_COLA (estado
    PENDIENTE) y devuelve el ID de cola. Commit propio para que la cola
    persista aunque el caller luego levante PeriodoFuturoEncolado."""
    fecha = (d.get('fecha') or '')[:10]
    ano_obj = int(fecha[:4]); mes_obj = int(fecha[5:7])
    _ncf = str(d.get('ncf') or '')
    with client.cursor() as cur:
        idvar = cur.var(int)
        cur.execute(
            "INSERT INTO CXP.TCXP_DOCUMENTO_COLA("
            " no_cia,punto,ano_objetivo,mes_objetivo,origen,tipo_docu,"
            " no_proveedor,ncf,fecha_doc,valor,payload,estado,usuario,fecha_encolado"
            ") VALUES(:1,:2,:3,:4,:5,:6,:7,:8,TO_DATE(:9,'YYYY-MM-DD'),:10,:11,"
            " 'PENDIENTE',:12,SYSDATE) RETURNING id INTO :13",
            [d['no_cia'], d.get('punto', '01'), ano_obj, mes_obj, origen,
             d.get('tipo_docu'), str(d.get('no_proveedor') or ''), _ncf[:30],
             fecha, float(d.get('valor_original') or d.get('valor') or 0),
             json.dumps(d, default=str), d.get('usuario', 'API'), idvar])
        cur.connection.commit()
        rid = idvar.getvalue()
        return int(rid[0] if isinstance(rid, list) else rid)


def _norm_fecha(valor, *, campo: str, requerido: bool = False):
    """Valida/normaliza una fecha 'YYYY-MM-DD' antes de pasarla a
    TO_DATE(:x,'YYYY-MM-DD'). El <input type="date"> de HTML permite girar el
    año a 5 dígitos (ej. 26810), lo que hace que TO_DATE reviente con
    ORA-01861 ('literal does not match format string') y el usuario solo ve
    ese error críptico (ver reporte de soporte de ACLASE, 2026-08-07:
    'ERROR AL ENTRAR UN DOCUMENTO EN CUENTA POR PAGAR'). Aquí lo convertimos
    en un mensaje claro y rechazamos años fuera de rango antes de tocar Oracle.
    """
    s = (str(valor or '')).strip()
    if not s:
        if requerido:
            raise ValueError(f'La {campo} es obligatoria.')
        return None
    # Solo aceptamos el formato ISO que emite <input type="date">.
    try:
        d = datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        raise ValueError(
            f'La {campo} "{s}" no es válida. Use el formato de fecha del '
            f'calendario (día/mes/año).')
    if not (1900 <= d.year <= 2999):
        raise ValueError(
            f'El año de la {campo} ({d.year}) está fuera de rango. '
            f'Revise que la fecha esté bien escrita.')
    return d.strftime('%Y-%m-%d')


def list_proveedores(search='', activo=''):
    # Busca por nombre, RNC o código. no_proveedor está almacenado con
    # padding a 6 dígitos ('000310'), así que "310" también debe encontrarlo
    # y el match exacto de código se ordena de primero.
    conditions = ['1=1']
    params = []
    s = (search or '').strip()
    if s:
        like = f'%{s.upper()}%'
        params += [like, like, like]
        conditions.append(
            f"(UPPER(p.nombre) LIKE :{len(params)-2} OR UPPER(p.rnc) LIKE :{len(params)-1}"
            f" OR TRIM(p.no_proveedor) LIKE :{len(params)})")
    if activo:
        params.append(activo)
        conditions.append(f"p.activo = :{len(params)}")
    order = 'p.nombre'
    if s:
        exacto = s.zfill(6) if s.isdigit() else s.upper()
        params.append(exacto)
        order = (f"CASE WHEN UPPER(TRIM(p.no_proveedor)) = :{len(params)} "
                 f"THEN 0 ELSE 1 END, p.nombre")
    where = ' AND '.join(conditions)
    sql = f"""
        SELECT p.no_proveedor, p.nombre, p.rnc, p.cedula,
               p.telefono, p.celular, p.e_mail,
               p.direccion, p.plazo_pago, p.activo,
               p.excento_itbis, p.categoria, p.clasificacion,
               p.cuenta_banco, p.codigo_banco, p.tipo_cuenta
        FROM CXP.TCXP_DPROVEEDOR p
        WHERE {where}
        ORDER BY {order}
    """
    return client.fetch_dicts(sql, params)


def get_proveedor(no_proveedor):
    """Busca un proveedor por su código.

    El usuario suele teclear `199` pero `TCXP_DPROVEEDOR.no_proveedor` está
    almacenado con padding a 6 dígitos (`000199`). Para entradas puramente
    numéricas probamos primero el valor literal y luego con LPAD a 6.
    """
    raw = (str(no_proveedor) if no_proveedor is not None else '').strip()
    candidates: list[str] = []
    if raw:
        candidates.append(raw)
        if raw.isdigit() and len(raw) < 6:
            candidates.append(raw.zfill(6))
        elif raw.isdigit() and len(raw) > 6:
            candidates.append(str(int(raw)))  # quita ceros a la izquierda
    sql = (
        "SELECT p.no_proveedor, p.nombre, p.rnc, p.cedula, "
        "p.telefono, p.celular, p.fax, p.e_mail, p.web_site, "
        "p.direccion, p.encargado, p.plazo_pago, p.activo, "
        "p.excento_itbis, p.categoria, p.clasificacion, "
        "p.cuenta_banco, p.codigo_banco, p.tipo_cuenta "
        "FROM CXP.TCXP_DPROVEEDOR p WHERE p.no_proveedor = :1"
    )
    seen: set[str] = set()
    for cand in candidates:
        if cand in seen:
            continue
        seen.add(cand)
        rows = client.fetch_dicts(sql, [cand])
        if rows:
            return rows[0]
    return None


def save_proveedor(data: dict):
    no_proveedor = (data.get('no_proveedor') or '').strip()
    with client.cursor() as cur:
        if no_proveedor:
            existing_rows = client.fetch_dicts(
                "SELECT * FROM CXP.TCXP_DPROVEEDOR WHERE no_proveedor=:1", [no_proveedor])
            ex = existing_rows[0] if existing_rows else {}
            cur.execute("""
                UPDATE CXP.TCXP_DPROVEEDOR SET
                    nombre=:1, rnc=:2, cedula=:3,
                    telefono=:4, celular=:5, fax=:6,
                    e_mail=:7, web_site=:8, direccion=:9,
                    encargado=:10, plazo_pago=:11, activo=:12,
                    excento_itbis=:13, categoria=:14, clasificacion=:15,
                    cuenta_banco=:16, codigo_banco=:17, tipo_cuenta=:18
                WHERE no_proveedor=:19
            """, [
                data.get('nombre') or ex.get('nombre', ''),
                data.get('rnc') if data.get('rnc') is not None else ex.get('rnc', ''),
                data.get('cedula') if data.get('cedula') is not None else ex.get('cedula', ''),
                data.get('telefono') if data.get('telefono') is not None else ex.get('telefono', ''),
                data.get('celular') if data.get('celular') is not None else ex.get('celular', ''),
                data.get('fax') if data.get('fax') is not None else ex.get('fax', ''),
                data.get('e_mail') if data.get('e_mail') is not None else ex.get('e_mail', ''),
                data.get('web_site') if data.get('web_site') is not None else ex.get('web_site', ''),
                data.get('direccion') if data.get('direccion') is not None else ex.get('direccion', ''),
                data.get('encargado') if data.get('encargado') is not None else ex.get('encargado', ''),
                data.get('plazo_pago') if data.get('plazo_pago') is not None else ex.get('plazo_pago', 0),
                data.get('activo') or ex.get('activo', 'S'),
                data.get('excento_itbis') or ex.get('excento_itbis', 'N'),
                data.get('categoria') if data.get('categoria') is not None else ex.get('categoria', ''),
                data.get('clasificacion') if data.get('clasificacion') is not None else ex.get('clasificacion', ''),
                data.get('cuenta_banco') if data.get('cuenta_banco') is not None else ex.get('cuenta_banco', ''),
                data.get('codigo_banco') if data.get('codigo_banco') is not None else ex.get('codigo_banco', ''),
                data.get('tipo_cuenta') if data.get('tipo_cuenta') is not None else ex.get('tipo_cuenta', ''),
                no_proveedor,
            ])
        else:
            cur.execute("""
                SELECT NVL(MAX(TO_NUMBER(no_proveedor)), 0) + 1
                FROM CXP.TCXP_DPROVEEDOR
                WHERE REGEXP_LIKE(no_proveedor, '^[0-9]+$')
            """, [])
            row = cur.fetchone()
            no_proveedor = str(int(row[0])).zfill(6)
            cur.execute("""
                INSERT INTO CXP.TCXP_DPROVEEDOR
                    (no_proveedor, nombre, rnc, cedula, telefono, celular, fax,
                     e_mail, web_site, direccion, encargado, plazo_pago, activo,
                     excento_itbis, categoria, clasificacion,
                     cuenta_banco, codigo_banco, tipo_cuenta)
                VALUES (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16,:17,:18,:19)
            """, [
                no_proveedor,
                data.get('nombre', ''), data.get('rnc', ''), data.get('cedula', ''),
                data.get('telefono', ''), data.get('celular', ''), data.get('fax', ''),
                data.get('e_mail', ''), data.get('web_site', ''), data.get('direccion', ''),
                data.get('encargado', ''), data.get('plazo_pago', 0), data.get('activo', 'S'),
                data.get('excento_itbis', 'N'), data.get('categoria', ''), data.get('clasificacion', ''),
                data.get('cuenta_banco', ''), data.get('codigo_banco', ''), data.get('tipo_cuenta', ''),
            ])
        cur.connection.commit()
    return {'no_proveedor': no_proveedor}


def list_documentos(no_cia, punto, no_proveedor='', tipo='', no_doc='',
                    desde='', hasta='', status='A', ncf=''):
    conditions = ['d.no_cia=:1', 'd.punto=:2']
    params = [no_cia, punto]
    if no_proveedor:
        params.append(no_proveedor)
        conditions.append(f'd.no_proveedor=:{len(params)}')
    if tipo:
        params.append(tipo)
        conditions.append(f'd.tipo_docu=:{len(params)}')
    if no_doc:
        params.append(f"%{no_doc}%")
        conditions.append(f'd.no_docu LIKE :{len(params)}')
    if ncf:
        # Compara contra el NCF DGI compuesto (prefijo + numero con ceros a
        # la izquierda, igual formula que composeNcfDgi en el frontend:
        # E-series usa 10 digitos, el resto 8) para que buscar "B0100003788"
        # o solo "3788" encuentre el documento igual que se ve en pantalla.
        params.append(f"%{ncf.strip().upper()}%")
        conditions.append(
            "UPPER(NVL(d.posiciones_fijas_ncf,'') || LPAD(TO_CHAR(NVL(d.ncf,0)), "
            "CASE WHEN UPPER(NVL(d.posiciones_fijas_ncf,'')) LIKE 'E%' THEN 10 ELSE 8 END, '0')) "
            f"LIKE :{len(params)}"
        )
    if desde:
        params.append(desde)
        conditions.append(f"d.fecha>=TO_DATE(:{len(params)},'YYYY-MM-DD')")
    if hasta:
        params.append(hasta)
        conditions.append(f"d.fecha<=TO_DATE(:{len(params)},'YYYY-MM-DD')")
    if status == 'A':
        # STATUS='A' en TCXP_DOCUMENTO significa "recien entrado/sin
        # contabilizar", NO "abierto/activo" -- la cartera pendiente real
        # vive casi toda en status='C' (contabilizado). Filtrar por
        # d.status='A' literal esconde el 97% de la deuda real y solo deja
        # ver residuos migrados con saldo 0. "Abiertos" en la UI = pendiente,
        # igual semantica ya aplicada en estado_cuenta/get_aging.
        conditions.append('NVL(d.saldo,0)<>0')
    elif status:
        params.append(status)
        conditions.append(f'd.status=:{len(params)}')
    where = ' AND '.join(conditions)
    sql = f"""
        SELECT d.no_cia, d.punto, d.tipo_docu, d.no_docu,
               d.no_proveedor, p.nombre AS nombre_proveedor,
               TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
               TO_CHAR(d.fecha_vence,'YYYY-MM-DD') AS fecha_vence,
               d.valor_original, d.saldo, d.status,
               d.ncf, d.posiciones_fijas_ncf, d.tipo_transaccion,
               d.impuesto, d.itbis_retenido, d.isr_retenido,
               d.tipo_movi, d.forma_pago, d.pago_bloqueado
        FROM CXP.TCXP_DOCUMENTO d
        JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor = d.no_proveedor
        WHERE {where}
        ORDER BY NVL(d.fecha_sysdate, d.fecha) DESC, d.no_docu DESC
    """
    return client.fetch_dicts(sql, params)


def get_documento(no_cia, punto, tipo_docu, no_docu):
    sql = """
        SELECT d.no_cia, d.punto, d.tipo_docu, d.no_docu,
               d.no_proveedor, p.nombre AS nombre_proveedor,
               p.direccion AS direccion_proveedor,
               p.telefono AS telefono_proveedor,
               p.celular  AS celular_proveedor,
               TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
               TO_CHAR(d.fecha_vence,'YYYY-MM-DD') AS fecha_vence,
               d.valor_original, d.saldo, d.status,
               d.ncf, d.posiciones_fijas_ncf, d.rnc,
               d.tipo_transaccion, d.impuesto,
               d.itbis_retenido, d.isr_retenido,
               d.tipo_movi, d.forma_pago, d.debito, d.credito,
               d.pago_bloqueado, d.detalle, d.usuario,
               d.tipo_gasto, d.tipo_retencion,
               d.valor_bienes, d.valor_servicio, d.isc, d.otros_impuestos, d.propina,
               d.tipo_docu_r, d.no_docu_r
        FROM CXP.TCXP_DOCUMENTO d
        JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor = d.no_proveedor
        WHERE d.no_cia=:1 AND d.punto=:2 AND d.tipo_docu=:3 AND d.no_docu=:4
    """
    sql_lineas = """
        SELECT dc.cuenta,
               dc.centro_costo,
               dc.no_componente,
               dc.monto,
               dc.tipo_movi
        FROM CXP.TCXP_DCDOCU dc
        WHERE dc.no_cia=:1 AND dc.punto=:2 AND dc.tipo_docu=:3 AND dc.no_docu=:4
        ORDER BY dc.no_componente NULLS FIRST, dc.cuenta
    """
    rows = client.fetch_dicts(sql, [no_cia, punto, tipo_docu, no_docu])
    if not rows:
        return None
    doc = rows[0]
    # NCF DGI compuesto (prefijo + LPAD(NCF,8,'0')) para que el frontend
    # muestre el comprobante completo.
    from .fat_repo import _compose_ncf_dgi
    doc['ncf_dgi'] = _compose_ncf_dgi(doc.get('posiciones_fijas_ncf'), doc.get('ncf'))
    doc['lineas'] = client.fetch_dicts(sql_lineas, [no_cia, punto, tipo_docu, no_docu])
    # Si es un credito (FP/FT/NC/AC), traer que debitos (ND/AD/BD) ya se le
    # aplicaron -- el grid "Doc(s) de Debito Afectado" que el legado
    # (FCXP501) muestra al consultar/editar una factura.
    if (doc.get('tipo_movi') or '').strip().upper() == 'C':
        doc['debitos_aplicados'] = get_debitos_aplicados(no_cia, punto, tipo_docu, no_docu)
    return doc


def estado_cuenta(no_cia: str, no_proveedor: str, punto: str = ''):
    """Estado de cuenta del proveedor: header con datos + tabla de documentos
    con saldo pendiente y envejecimiento por edad de la fecha de emisión.

    Pendiente = saldo<>0 sin importar status: 'A' es el documento recién
    entrado y 'C' el ya actualizado/contabilizado — ambos pueden deber.
    """
    raw = (str(no_proveedor) if no_proveedor is not None else '').strip()
    if raw.isdigit() and len(raw) < 6:
        raw = raw.zfill(6)
    prov = client.fetch_dicts(
        "SELECT no_proveedor, nombre, "
        "       NVL(rnc,'') rnc, NVL(direccion,'') direccion, "
        "       NVL(telefono,'') telefono, NVL(celular,'') celular, "
        "       NVL(e_mail,'') email, NVL(encargado,'') encargado, "
        "       NVL(plazo_pago,0) dias "
        "FROM CXP.TCXP_DPROVEEDOR WHERE no_proveedor=:1",
        [raw])
    if not prov:
        return None
    no_proveedor = str(prov[0]['no_proveedor'])

    where = "WHERE d.no_cia=:1 AND d.no_proveedor=:2 AND NVL(d.saldo,0)<>0"
    params = [no_cia, no_proveedor]
    if punto:
        where += " AND d.punto=:3"
        params.append(punto)
    docs = client.fetch_dicts(
        "SELECT d.no_docu no_doc, d.tipo_docu tipo_doc, d.punto, "
        "       TO_CHAR(d.fecha,'YYYY-MM-DD') fecha, "
        "       TO_CHAR(d.fecha_vence,'YYYY-MM-DD') fecha_vence, "
        "       NVL(d.valor_original,0) valor, NVL(d.saldo,0) saldo, "
        "       d.ncf, d.posiciones_fijas_ncf, d.detalle, "
        "       TRUNC(SYSDATE)-TRUNC(d.fecha) dias_vencido, "
        "       NVL(t.tipo_movi,'D') tipo_movi, "
        "       NVL(t.descri, d.tipo_docu) tipo_label "
        "FROM CXP.TCXP_DOCUMENTO d "
        "LEFT JOIN CXP.TCXP_TDOCU t ON t.tipo_docu=d.tipo_docu "
        f"{where} ORDER BY d.fecha, d.no_docu",
        params)
    # Componer NCF DGI: prefijo (B01..B15) + LPAD(NCF,8,'0') para que el
    # frontend muestre el comprobante completo y no solo el correlativo.
    from .fat_repo import _compose_ncf_dgi
    for d in docs:
        d['ncf'] = _compose_ncf_dgi(d.get('posiciones_fijas_ncf'), d.get('ncf')) or d.get('ncf')

    aging = {'d_0_30': 0.0, 'd_31_60': 0.0, 'd_61_90': 0.0, 'd_mas_90': 0.0}
    total_debito = 0.0
    total_credito = 0.0
    for d in docs:
        sld = float(d.get('saldo') or 0)
        dv = int(d.get('dias_vencido') or 0)
        if sld > 0:
            total_debito += sld
            if dv <= 30: aging['d_0_30'] += sld
            elif dv <= 60: aging['d_31_60'] += sld
            elif dv <= 90: aging['d_61_90'] += sld
            else: aging['d_mas_90'] += sld
        elif sld < 0:
            total_credito += abs(sld)
    total_pendiente = total_debito - total_credito

    row = prov[0]
    return {
        'proveedor': {
            'no_proveedor': str(row['no_proveedor']).strip(),
            'nombre': row['nombre'],
            'rnc': row['rnc'], 'direccion': row['direccion'],
            'telefono': row['telefono'] or row['celular'],
            'email': row['email'],
            'encargado': row['encargado'],
            'dias': int(row['dias'] or 0),
        },
        'total_pendiente': total_pendiente,
        'total_debito': total_debito,
        'total_credito': total_credito,
        'aging': aging,
        'documentos': docs,
    }


def get_documentos_afectados(no_cia, punto, tipo_docu, no_docu):
    """TCXP_REFEDOCU + datos del documento referenciado (saldo actual).
    Direccion: dado un debito (ND/AD/BD), que creditos (facturas/FP) afecto."""
    sql = """
        SELECT r.tipo_refe AS tipo_doc, r.no_refe AS no_doc,
               r.no_cuota, r.monto,
               TO_CHAR(rd.fecha,'YYYY-MM-DD') AS fecha,
               NVL(rd.valor_original,0) AS valor_original,
               NVL(rd.saldo,0) AS saldo
        FROM CXP.TCXP_REFEDOCU r
        LEFT JOIN CXP.TCXP_DOCUMENTO rd
          ON rd.no_cia=r.no_cia AND rd.punto=r.punto
         AND rd.tipo_docu=r.tipo_refe AND rd.no_docu=r.no_refe
        WHERE r.no_cia=:1 AND r.punto=:2 AND r.tipo_docu=:3 AND r.no_docu=:4
        ORDER BY r.no_cuota, r.tipo_refe, r.no_refe
    """
    return client.fetch_dicts(sql, [no_cia, punto, tipo_docu, no_docu])


def get_debitos_aplicados(no_cia, punto, tipo_docu, no_docu):
    """TCXP_REFEDOCU en sentido inverso: dado un credito (FP/FT/NC/AC), que
    debitos (ND/AD/BD) se le aplicaron -- equivale al grid "Doc(s) de Debito
    Afectado" que el legado (FCXP501) muestra al consultar una factura."""
    sql = """
        SELECT r.tipo_docu AS tipo_doc, r.no_docu AS no_doc,
               r.no_cuota, r.monto,
               TO_CHAR(rd.fecha,'YYYY-MM-DD') AS fecha,
               NVL(rd.valor_original,0) AS valor_original,
               NVL(rd.saldo,0) AS saldo
        FROM CXP.TCXP_REFEDOCU r
        LEFT JOIN CXP.TCXP_DOCUMENTO rd
          ON rd.no_cia=r.no_cia AND rd.punto=r.punto
         AND rd.tipo_docu=r.tipo_docu AND rd.no_docu=r.no_docu
        WHERE r.no_cia=:1 AND r.punto=:2 AND r.tipo_refe=:3 AND r.no_refe=:4
        ORDER BY r.no_cuota, r.tipo_docu, r.no_docu
    """
    return client.fetch_dicts(sql, [no_cia, punto, tipo_docu, no_docu])


def list_tipos_docu():
    return client.fetch_dicts(
        "SELECT tipo_docu AS codigo, descri AS nombre, tipo_movi FROM CXP.TCXP_TDOCU ORDER BY tipo_docu", [])


def get_aging(no_cia, punto='', no_proveedor=''):
    # Pendiente = saldo<>0; el status 'A'/'C' (entrado/actualizado) no
    # determina si el documento debe — filtrar por 'A' escondía la cartera.
    conditions = ["d.no_cia=:1", "NVL(d.saldo,0)<>0"]
    params = [no_cia]
    if punto:
        params.append(punto)
        conditions.append(f'd.punto=:{len(params)}')
    if no_proveedor:
        params.append(no_proveedor)
        conditions.append(f'd.no_proveedor=:{len(params)}')
    where = ' AND '.join(conditions)
    sql = f"""
        SELECT d.no_proveedor, p.nombre,
               SUM(CASE WHEN d.fecha_vence IS NULL OR SYSDATE - d.fecha_vence <= 0
                        THEN d.saldo ELSE 0 END) AS corriente,
               SUM(CASE WHEN SYSDATE - d.fecha_vence BETWEEN 1  AND 30  THEN d.saldo ELSE 0 END) AS d30,
               SUM(CASE WHEN SYSDATE - d.fecha_vence BETWEEN 31 AND 60  THEN d.saldo ELSE 0 END) AS d60,
               SUM(CASE WHEN SYSDATE - d.fecha_vence BETWEEN 61 AND 90  THEN d.saldo ELSE 0 END) AS d90,
               SUM(CASE WHEN SYSDATE - d.fecha_vence > 90               THEN d.saldo ELSE 0 END) AS mas90,
               SUM(d.saldo) AS total
        FROM CXP.TCXP_DOCUMENTO d
        JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor = d.no_proveedor
        WHERE {where}
        GROUP BY d.no_proveedor, p.nombre
        ORDER BY p.nombre
    """
    return client.fetch_dicts(sql, params)


def list_tipos_gasto():
    return client.fetch_dicts(
        "SELECT tipo_gasto, descripcion FROM CXP.TCXP_TCOSTO_GASTO "
        "ORDER BY tipo_gasto", [])


def list_tipos_retencion():
    return client.fetch_dicts(
        "SELECT tipo_retencion, descripcion, por_defecto "
        "FROM CXP.TCXP_TIPO_RETENCION_DGII ORDER BY tipo_retencion", [])


def list_formas_pago():
    return client.fetch_dicts(
        "SELECT forma_pago, descripcion, por_defecto "
        "FROM CXP.TCXP_FORMA_PAGO_DGII ORDER BY forma_pago", [])


def get_proveedor_ncf_info(no_cia, punto, no_proveedor):
    """Devuelve la configuracion NCF asignada al proveedor para (cia, punto).

    Si TCXP_BPROVEEDOR.codigo_ncf esta NULL → proveedor formal (operador
    digita el NCF que viene en la factura). Si esta lleno (PI-001, etc.)
    → proveedor informal: el sistema toma POSICIONES_FIJAS (B11) y
    PROX_NCF de CNT.TCNT_NCF y los pre-llena en el form.
    """
    rows = client.fetch_dicts(
        "SELECT bp.codigo_ncf, n.posiciones_fijas, n.prox_ncf, "
        "n.ncf_inicial, n.ncf_final, n.descripcion, n.ncf_manual, "
        "n.ncf_opcional "
        "FROM CXP.TCXP_BPROVEEDOR bp "
        "LEFT JOIN CNT.TCNT_NCF n ON n.codigo_ncf = bp.codigo_ncf "
        "WHERE bp.no_cia=:1 AND bp.punto=:2 AND bp.no_proveedor=:3",
        [no_cia, punto, no_proveedor])
    if not rows:
        return {"codigo_ncf": None}
    r = rows[0]
    if not r.get('codigo_ncf'):
        return {"codigo_ncf": None}
    return r


def get_proveedor_cuenta(no_cia, punto, no_proveedor):
    """Account summary for FCXP502/503 header, y cuenta/cuenta_prima por
    defecto para el movimiento contable de Entrada de Documentos (Fcxp201):
    TCXP_TPROVEEDOR.CUENTA/CUENTA_PRIMA/CENTRO_COSTO estan configuradas por
    TIPO_PROVEEDOR (ej. 01=Proveedores Nacionales -> cuenta 2101-01,
    Cuentas por Pagar), no por proveedor individual."""
    prov = client.fetch_dicts(
        "SELECT d.no_proveedor, d.nombre, d.rnc, d.categoria, d.clasificacion, "
        "d.tipo_proveedor, d.cuenta_gasto, t.cuenta, t.cuenta_prima, t.centro_costo "
        "FROM CXP.TCXP_DPROVEEDOR d "
        "LEFT JOIN CXP.TCXP_TPROVEEDOR t ON t.tipo_proveedor=d.tipo_proveedor "
        "WHERE d.no_proveedor=:1",
        [no_proveedor])
    if not prov:
        return None
    data = prov[0]
    fin = client.fetch_dicts("""
        SELECT
            NVL(SUM(saldo), 0) AS balance,
            NVL(SUM(CASE WHEN tipo_movi='Credito' THEN valor_original ELSE 0 END), 0) AS compras_acumuladas,
            NVL(SUM(CASE WHEN tipo_movi='Debito'  THEN valor_original ELSE 0 END), 0) AS pagos_acumulados,
            MAX(CASE WHEN tipo_movi='Credito' THEN TO_CHAR(fecha,'YYYY-MM-DD') END) AS fecha_ultima_compra,
            MAX(CASE WHEN tipo_movi='Debito'  THEN TO_CHAR(fecha,'YYYY-MM-DD') END) AS fecha_ultimo_pago
        FROM CXP.TCXP_DOCUMENTO
        WHERE no_cia=:1 AND punto=:2 AND no_proveedor=:3
    """, [no_cia, punto, no_proveedor])
    data.update(fin[0] if fin else {})
    return data


def get_cuenta_itbis_default(no_cia: str) -> str:
    """Cuenta contable de ITBIS deducible/pagado mas usada historicamente
    por esta empresa en TCXP_DCDOCU (no hay columna de config dedicada
    para esto, a diferencia de CUENTA_ITBIS_RETENIDO/CUENTA_ISR en
    TCNT_CIAS). Comprobado contra datos reales: 2106-02 domina en las
    companias 01-04 (miles de lineas), la 05 usa 1304-05."""
    row = client.fetch_one(
        "SELECT cuenta FROM ("
        "  SELECT dc.cuenta FROM CXP.TCXP_DCDOCU dc "
        "  JOIN CNT.TCNT_CATALOGO c ON c.cuenta=dc.cuenta "
        "  WHERE dc.no_cia=:1 AND UPPER(c.nombre) LIKE '%ITBIS%' "
        "  AND UPPER(c.nombre) NOT LIKE '%RETEN%' "
        "  GROUP BY dc.cuenta ORDER BY COUNT(*) DESC"
        ") WHERE ROWNUM = 1",
        [no_cia])
    return row[0] if row else ''


def get_cuenta_compra_default(no_cia: str) -> str:
    """Cuenta de compras/gasto mas usada como DEBITO en TCXP_DCDOCU para
    esta compania, excluyendo cuentas 21xx (pasivos, saldo por pagar) y
    la cuenta de ITBIS deducible. Se usa como fallback cuando un
    proveedor no tiene TCXP_DPROVEEDOR.CUENTA_GASTO configurada."""
    row = client.fetch_one(
        "SELECT cuenta FROM ("
        "  SELECT dc.cuenta FROM CXP.TCXP_DCDOCU dc "
        "  WHERE dc.no_cia=:1 AND dc.tipo_movi='D' "
        "  AND dc.cuenta NOT LIKE '21%' "
        "  AND dc.cuenta NOT LIKE '2106%' "
        "  AND dc.cuenta IS NOT NULL "
        "  GROUP BY dc.cuenta ORDER BY COUNT(*) DESC"
        ") WHERE ROWNUM = 1",
        [no_cia])
    return row[0] if row else ''


def list_cuentas_proveedor(no_cia, punto, no_proveedor, tipo_movi='', en_cero='N'):
    """FCXP502: document list for one proveedor"""
    conditions = ['d.no_cia=:1', 'd.punto=:2', 'd.no_proveedor=:3']
    params = [no_cia, punto, no_proveedor]
    if tipo_movi:
        params.append(tipo_movi)
        conditions.append(f'd.tipo_movi=:{len(params)}')
    if en_cero == 'N':
        conditions.append('NVL(d.saldo,0) != 0')
    elif en_cero == 'S':
        conditions.append('NVL(d.saldo,0) = 0')
    where = ' AND '.join(conditions)
    return client.fetch_dicts(f"""
        SELECT d.tipo_docu, d.no_docu, d.tipo_movi,
               TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
               TO_CHAR(d.fecha_vence,'YYYY-MM-DD') AS fecha_vence,
               d.valor_original, NVL(d.saldo, 0) AS saldo
        FROM CXP.TCXP_DOCUMENTO d
        WHERE {where}
        ORDER BY d.fecha DESC, d.no_docu DESC
    """, params)


def list_movimientos_proveedor(no_cia, punto, no_proveedor, desde='', hasta=''):
    """FCXP503: movements for one proveedor ordered by date"""
    conditions = ['d.no_cia=:1', 'd.punto=:2', 'd.no_proveedor=:3']
    params = [no_cia, punto, no_proveedor]
    if desde:
        params.append(desde)
        conditions.append(f"d.fecha>=TO_DATE(:{len(params)},'YYYY-MM-DD')")
    if hasta:
        params.append(hasta)
        conditions.append(f"d.fecha<=TO_DATE(:{len(params)},'YYYY-MM-DD')")
    where = ' AND '.join(conditions)
    return client.fetch_dicts(f"""
        SELECT d.tipo_docu, d.no_docu, d.tipo_movi,
               TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
               d.valor_original,
               CASE WHEN d.tipo_movi='Debito'  THEN d.valor_original ELSE 0 END AS debito,
               CASE WHEN d.tipo_movi='Credito' THEN d.valor_original ELSE 0 END AS credito
        FROM CXP.TCXP_DOCUMENTO d
        WHERE {where}
        ORDER BY d.fecha, d.no_docu
    """, params)


# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

def list_cias():
    return client.fetch_dicts(
        "SELECT no_cia, descri AS descripcion, NVL(activa,'S') activa "
        "FROM CXP.TCXP_CIAS ORDER BY no_cia", [])


def list_puntos(no_cia: str):
    return client.fetch_dicts(
        "SELECT no_cia, punto, descri AS descripcion, NVL(activo,'S') activo, "
        "NVL(mes_proceso,1) mes_proceso, NVL(ano_proceso,2025) ano_proceso "
        "FROM CXP.TCXP_PUNTO WHERE no_cia=:1 ORDER BY punto",
        [no_cia])


def list_tproveedores():
    return client.fetch_dicts(
        "SELECT tipo_proveedor AS codigo, NVL(descri,'') descripcion, "
        "clase_proveedor, NVL(cuenta,'') cuenta, "
        "NVL(cuenta_prima,'') cuenta_prima, NVL(centro_costo,'') centro_costo "
        "FROM CXP.TCXP_TPROVEEDOR ORDER BY tipo_proveedor", [])


def list_tdocu_config():
    """Full tdocu with all columns for configuration screen."""
    return client.fetch_dicts(
        "SELECT tipo_docu, tipo_movi, NVL(descri,'') descripcion, "
        "tipo_transaccion, NVL(cuenta,'') cuenta, NVL(centro_costo,'') centro_costo "
        "FROM CXP.TCXP_TDOCU ORDER BY tipo_docu", [])


def list_ciudades():
    """Ciudades from shared CXC table (no CXP-specific copy exists)."""
    return client.fetch_dicts(
        "SELECT ciudad, descripcion FROM CXC.TCXC_CIUDAD ORDER BY ciudad", [])


def list_usuarios(no_cia: str = '', punto: str = ''):
    """Acceso de usuarios al módulo CxP (FCXP103)."""
    conditions = ['1=1']
    params = []
    if no_cia:
        params.append(no_cia)
        conditions.append('no_cia=:' + str(len(params)))
    if punto:
        params.append(punto)
        conditions.append('punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    return client.fetch_dicts(
        "SELECT no_cia, punto, usuario, "
        "NVL(por_defecto,'N') por_defecto, NVL(activo,'S') activo, "
        "NVL(hacer_transacciones,'N') hacer_transacciones, "
        "NVL(generar_listado_cxp,'N') generar_listado_cxp, "
        "NVL(crear_proveedor,'N') crear_proveedor, "
        "NVL(hacer_cierre,'N') hacer_cierre, "
        "NVL(asignar_proveedor,'N') asignar_proveedor, "
        "NVL(asignar_cuenta_bancaria,'N') asignar_cuenta_bancaria, "
        "NVL(liberar_debito,'N') liberar_debito, "
        "NVL(bloquear_pago,'N') bloquear_pago "
        "FROM CXP.TCXP_USUARIO WHERE " + where + " ORDER BY no_cia, punto, usuario",
        params)


_CXP_USUARIO_FLAGS = (
    'por_defecto', 'activo', 'hacer_transacciones', 'generar_listado_cxp',
    'crear_proveedor', 'hacer_cierre', 'asignar_proveedor',
    'asignar_cuenta_bancaria', 'liberar_debito', 'bloquear_pago',
)


def _coerce_sn(v) -> str:
    s = (str(v) if v is not None else '').strip().upper()
    return 'S' if s in ('S', 'Y', '1', 'TRUE') else 'N'


def upsert_usuario(data: dict) -> dict:
    """Crea o actualiza un acceso TCXP_USUARIO (PK: no_cia, punto, usuario)."""
    no_cia  = (data.get('no_cia') or '').strip()
    punto   = (data.get('punto') or '').strip()
    usuario = (data.get('usuario') or '').strip().upper()
    if not no_cia or not punto or not usuario:
        raise ValueError('no_cia, punto y usuario son requeridos')
    flags = {k: _coerce_sn(data.get(k)) for k in _CXP_USUARIO_FLAGS}
    with client.connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM CXP.TCXP_USUARIO WHERE no_cia=:1 AND punto=:2 AND usuario=:3",
            [no_cia, punto, usuario],
        )
        exists = cur.fetchone() is not None
        if exists:
            cur.execute(
                "UPDATE CXP.TCXP_USUARIO SET "
                "por_defecto=:1, activo=:2, hacer_transacciones=:3, "
                "generar_listado_cxp=:4, crear_proveedor=:5, hacer_cierre=:6, "
                "asignar_proveedor=:7, asignar_cuenta_bancaria=:8, "
                "liberar_debito=:9, bloquear_pago=:10 "
                "WHERE no_cia=:11 AND punto=:12 AND usuario=:13",
                [flags['por_defecto'], flags['activo'], flags['hacer_transacciones'],
                 flags['generar_listado_cxp'], flags['crear_proveedor'], flags['hacer_cierre'],
                 flags['asignar_proveedor'], flags['asignar_cuenta_bancaria'],
                 flags['liberar_debito'], flags['bloquear_pago'],
                 no_cia, punto, usuario],
            )
            action = 'updated'
        else:
            cur.execute(
                "INSERT INTO CXP.TCXP_USUARIO ("
                "no_cia, punto, usuario, por_defecto, activo, hacer_transacciones, "
                "generar_listado_cxp, crear_proveedor, hacer_cierre, asignar_proveedor, "
                "asignar_cuenta_bancaria, liberar_debito, bloquear_pago) VALUES ("
                ":1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13)",
                [no_cia, punto, usuario,
                 flags['por_defecto'], flags['activo'], flags['hacer_transacciones'],
                 flags['generar_listado_cxp'], flags['crear_proveedor'], flags['hacer_cierre'],
                 flags['asignar_proveedor'], flags['asignar_cuenta_bancaria'],
                 flags['liberar_debito'], flags['bloquear_pago']],
            )
            action = 'created'
        conn.commit()
    return {'ok': True, 'action': action,
            'no_cia': no_cia, 'punto': punto, 'usuario': usuario}


def delete_usuario(no_cia: str, punto: str, usuario: str) -> dict:
    if not no_cia or not punto or not usuario:
        raise ValueError('no_cia, punto y usuario son requeridos')
    with client.connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM CXP.TCXP_USUARIO WHERE no_cia=:1 AND punto=:2 AND usuario=:3",
            [no_cia, punto, usuario.upper()],
        )
        affected = cur.rowcount
        conn.commit()
    return {'ok': True, 'deleted': affected}


def rep_cuadre(no_cia: str, punto: str, mes: int, ano: int):
    """RCXP105 — Cuadre Contable por cuenta. Agrupa TCXP_DOCUMENTO del
    periodo por cuenta contable, sumando debe/haber según tipo_movi."""
    sql = (
        "SELECT cuenta, "
        "SUM(CASE WHEN tipo_movi = 'D' THEN valor_original ELSE 0 END) AS debe, "
        "SUM(CASE WHEN tipo_movi = 'C' THEN valor_original ELSE 0 END) AS haber, "
        "COUNT(*) AS docs "
        "FROM CXP.TCXP_DOCUMENTO "
        "WHERE no_cia = :1 AND punto = :2 "
        "AND EXTRACT(MONTH FROM fecha) = :3 AND EXTRACT(YEAR FROM fecha) = :4 "
        "AND NVL(status,'A') <> 'R' "
        "AND cuenta IS NOT NULL "
        "GROUP BY cuenta ORDER BY cuenta"
    )
    rows = client.fetch_dicts(sql, [no_cia, punto, mes, ano])
    total_debe = sum(float(r.get('debe') or 0) for r in rows)
    total_haber = sum(float(r.get('haber') or 0) for r in rows)
    return {
        'items': rows,
        'total_debe': total_debe,
        'total_haber': total_haber,
        'diferencia': total_debe - total_haber,
    }


def rep_retenciones(no_cia: str, punto: str, ano: int, no_proveedor: str = ''):
    """RCXP108 — Certificado de retenciones por proveedor en un año.
    Devuelve por proveedor el detalle de docs con ITBIS/ISR retenido."""
    conditions = [
        "d.no_cia = :1", "d.punto = :2",
        "EXTRACT(YEAR FROM d.fecha) = :3",
        "(NVL(d.itbis_retenido,0) > 0 OR NVL(d.isr_retenido,0) > 0)",
        "NVL(d.status,'A') <> 'R'",
    ]
    params = [no_cia, punto, ano]
    if no_proveedor:
        params.append(no_proveedor.strip())
        conditions.append('d.no_proveedor = :' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT d.no_proveedor, p.nombre AS nombre_proveedor, p.rnc AS rnc_proveedor, "
        "d.tipo_docu, d.no_docu, d.fecha, d.ncf, d.posiciones_fijas_ncf, "
        "d.valor_original, d.impuesto, NVL(d.itbis_retenido,0) AS itbis_retenido, "
        "NVL(d.isr_retenido,0) AS isr_retenido, d.tipo_retencion "
        "FROM CXP.TCXP_DOCUMENTO d "
        "LEFT JOIN CXP.TCXP_BPROVEEDOR b ON b.no_cia = d.no_cia AND b.punto = d.punto AND b.no_proveedor = d.no_proveedor "
        "LEFT JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor = d.no_proveedor "
        f"WHERE {where} "
        "ORDER BY d.no_proveedor, d.fecha"
    )
    rows = client.fetch_dicts(sql, params)
    # Agrupar por proveedor
    grouped: dict[str, dict] = {}
    for r in rows:
        key = (r.get('no_proveedor') or '').strip()
        g = grouped.setdefault(key, {
            'no_proveedor': key,
            'nombre_proveedor': r.get('nombre_proveedor') or '',
            'rnc_proveedor': r.get('rnc_proveedor') or '',
            'documentos': [],
            'total_itbis': 0.0,
            'total_isr': 0.0,
        })
        g['documentos'].append(r)
        g['total_itbis'] += float(r.get('itbis_retenido') or 0)
        g['total_isr'] += float(r.get('isr_retenido') or 0)
    total_itbis = sum(g['total_itbis'] for g in grouped.values())
    total_isr = sum(g['total_isr'] for g in grouped.values())
    return {
        'proveedores': list(grouped.values()),
        'total_itbis': total_itbis,
        'total_isr': total_isr,
        'count_docs': len(rows),
    }


def list_barrios(ciudad: str = ''):
    conditions = ['1=1']
    params = []
    if ciudad:
        params.append(ciudad)
        conditions.append('ciudad=:1')
    where = ' AND '.join(conditions)
    return client.fetch_dicts(
        "SELECT ciudad, barrio, descripcion FROM CXC.TCXC_BARRIO WHERE " + where +
        " ORDER BY ciudad, barrio", params)


# ─── REPORTES READ-ONLY ───────────────────────────────────────────────────────

def rep_alfabetico(no_cia: str = '', punto: str = '', search: str = ''):
    """Listado alfabetico de proveedores activos con saldo."""
    conditions = ["p.activo='S'"]
    params = []
    if search:
        params.append('%' + search.upper() + '%')
        conditions.append('UPPER(p.nombre) LIKE :' + str(len(params)))
    where = ' AND '.join(conditions)
    if no_cia and punto:
        n_cia_idx = len(params) + 1
        n_pun_idx = n_cia_idx + 1
        params += [no_cia, punto]
        sql = (
            "SELECT p.no_proveedor, p.nombre, p.rnc, p.telefono, "
            "p.e_mail, p.encargado, "
            "NVL(b.saldo_inicial,0) saldo_inicial, "
            "NVL(b.compras_acumuladas,0) compras, "
            "NVL(b.pagos_acumulados,0) pagos, "
            "NVL(b.saldo_inicial,0)+NVL(b.creditos,0)-NVL(b.debitos,0) saldo "
            "FROM CXP.TCXP_DPROVEEDOR p "
            "LEFT JOIN CXP.TCXP_BPROVEEDOR b "
            "  ON b.no_proveedor=p.no_proveedor "
            "  AND b.no_cia=:" + str(n_cia_idx) +
            "  AND b.punto=:" + str(n_pun_idx) +
            " WHERE " + where + " ORDER BY p.nombre"
        )
    else:
        sql = (
            "SELECT p.no_proveedor, p.nombre, p.rnc, p.telefono, "
            "p.e_mail, p.encargado, "
            "0 saldo_inicial, 0 compras, 0 pagos, 0 saldo "
            "FROM CXP.TCXP_DPROVEEDOR p "
            "WHERE " + where + " ORDER BY p.nombre"
        )
    rows = client.fetch_dicts(sql, params)
    return {'items': rows, 'count': len(rows)}


def rep_mayor_auxiliar(no_cia: str, punto: str, desde: str, hasta: str,
                       no_proveedor: str = ''):
    """Mayor auxiliar CxP: documentos por proveedor en rango de fechas."""
    conditions = [
        "d.no_cia=:1", "d.punto=:2",
        "d.fecha BETWEEN TO_DATE(:3,'YYYY-MM-DD') AND TO_DATE(:4,'YYYY-MM-DD')"
    ]
    params = [no_cia, punto, desde, hasta]
    if no_proveedor:
        params.append(no_proveedor)
        conditions.append('d.no_proveedor=:' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT d.no_proveedor, p.nombre AS nombre_proveedor, "
        "d.tipo_docu, d.no_docu, "
        "TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha, "
        "TO_CHAR(d.fecha_vence,'YYYY-MM-DD') AS fecha_vence, "
        "d.tipo_movi, d.valor_original, "
        "CASE WHEN d.tipo_movi='Credito' THEN d.valor_original ELSE 0 END credito, "
        "CASE WHEN d.tipo_movi='Debito'  THEN d.valor_original ELSE 0 END debito, "
        "NVL(d.saldo,0) saldo, d.status, d.ncf, d.detalle "
        "FROM CXP.TCXP_DOCUMENTO d "
        "JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor=d.no_proveedor "
        "WHERE " + where + " ORDER BY d.no_proveedor, d.fecha, d.no_docu"
    )
    rows = client.fetch_dicts(sql, params)
    total_credito = sum((r.get('credito') or 0) for r in rows)
    total_debito  = sum((r.get('debito')  or 0) for r in rows)
    return {'items': rows, 'total_credito': total_credito, 'total_debito': total_debito}


# Catalogo DGII "Tipo de Bienes y Servicios Comprados" (formato 606). El
# reporte del legado imprime "NN-DESCRIPCION" (ej. "02-GASTOS POR TRABAJO,
# SUMINISTROS Y SERVICIOS"). Se replica textual para casar el Excel del 606.
_TIPO_GASTO_606 = {
    '01': 'GASTOS DE PERSONAL',
    '02': 'GASTOS POR TRABAJO, SUMINISTROS Y SERVICIOS',
    '03': 'ARRENDAMIENTOS',
    '04': 'GASTOS DE ACTIVOS FIJOS',
    '05': 'GASTOS DE REPRESENTACION',
    '06': 'OTRAS DEDUCCIONES ADMITIDAS',
    '07': 'GASTOS FINANCIEROS',
    '08': 'GASTOS EXTRAORDINARIOS',
    '09': 'COMPRAS Y GASTOS QUE FORMARAN PARTE DEL COSTO DE VENTA',
    '10': 'ADQUISICIONES DE ACTIVOS',
    '11': 'GASTOS DE SEGUROS',
}


def _tg2(v) -> str:
    """Normaliza tipo_gasto a string de 2 digitos ('02'); '' si no aplica."""
    s = str(v or '').strip()
    return f"{int(s):02d}" if s.isdigit() else ''


def _tipo_gasto_label(v) -> str:
    """'02-GASTOS POR TRABAJO, SUMINISTROS Y SERVICIOS' o '' — igual que el 606."""
    tg = _tg2(v)
    return f"{tg}-{_TIPO_GASTO_606[tg]}" if tg in _TIPO_GASTO_606 else ''


def _base_606(r) -> float:
    """Monto Facturado del 606 = base de bienes/servicios SIN impuestos.

    valor_original NO es el bruto: es el NETO A PAGAR al proveedor =
      base + ITBIS + ISC + otros_impuestos + propina - itbis_retenido - isr_retenido
    (las retenciones se le descuentan al proveedor y las paga la empresa a la
    DGII; ISC/otros/propina se le pagan pero no son parte del monto facturado).
    Despejando la base:
      base = valor_original - impuesto - isc - otros_impuestos - propina
             + itbis_retenido + isr_retenido
    Verificado contra el 606 real de julio/2026 (legado): RMM FP-0008546
    14638-2340+702 = 13000; TC BANCO POPULAR FP-0008548 840-116.31-64.62-12.92
    = 646.15 (exacto). Para docs sin retencion/isc/otros/propina el resultado
    es identico a valor_original - impuesto (no cambia lo que ya estaba bien).
    """
    g = lambda k: float(r.get(k) or 0)
    return (g('valor_original') - g('impuesto') - g('isc') - g('otros_impuestos')
            - g('propina') + g('itbis_retenido') + g('isr_retenido'))


def _rows_acc_606(no_cia: str, anio: int, mes: int, punto: str = ''):
    """Comprobantes de Caja Chica (ACC) con NCF real -- deben reportarse en
    el 606 igual que las facturas de CXP. Antes de esto el 606 (pantalla,
    Excel y archivo DGII) solo leia CXP.TCXP_DOCUMENTO: los comprobantes de
    caja chica con NCF nunca aparecian en ningun lado (reportado 2026-08-12
    por Pilar/JC via ticket de soporte -- ver 606 reporte / foto del error).
    Mismas reglas de exclusion que rep_606/archivo_dgii_606: sin NCF no
    cuenta, B02/E32 no es comprobante de compra, anulado no es compra real.
    Devuelve dicts con ambos alias de clave ('rnc'/'rnc_proveedor', etc.)
    para servir tanto a rep_606 (usa *_proveedor) como a archivo_dgii_606
    (usa las claves crudas sin alias).
    """
    conditions = [
        "a.no_cia=:1", "EXTRACT(YEAR FROM a.fecha)=:2", "EXTRACT(MONTH FROM a.fecha)=:3",
        "TRIM(a.ncf) IS NOT NULL",
        "NVL(UPPER(a.posiciones_fijas_ncf),'') NOT IN ('B02','E32')",
        "NVL(a.anulado,'N') <> 'S'",
    ]
    params = [no_cia, anio, mes]
    if punto:
        params.append(punto)
        conditions.append('a.punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT a.rnc, b.nombre AS nombre_proveedor, "
        "a.ncf, a.posiciones_fijas_ncf, "
        "TO_CHAR(a.fecha,'YYYY-MM-DD') AS fecha, "
        "TO_CHAR(a.fecha_vence_ncf,'YYYY-MM-DD') AS fecha_vence, "
        "a.no_cia, a.punto, a.detalle, a.no_docu, a.no_bene, a.no_reposicion, "
        "a.tipo_gasto_dgii, "
        "NVL(a.valor,0) valor_original, NVL(a.impuesto,0) impuesto, "
        "NVL(a.isc,0) isc, NVL(a.otros_impuestos,0) otros_impuestos, "
        "NVL(a.propina,0) propina, "
        "NVL(a.valor_bienes,0) valor_bienes, NVL(a.valor_servicio,0) valor_servicio, "
        "NVL(a.forma_pago,1) forma_pago "
        "FROM ACC.TACC_DOCUMENTO a "
        "LEFT JOIN ACC.TACC_BENEFICIARIO b ON b.no_bene=a.no_bene "
        "WHERE " + where + " ORDER BY a.fecha, a.ncf"
    )
    raw = client.fetch_dicts(sql, params)
    rows = []
    for a in raw:
        rows.append({
            'rnc': a.get('rnc'), 'rnc_proveedor': a.get('rnc'),
            'nombre_proveedor': a.get('nombre_proveedor'),
            'tipo_transaccion': '',
            'ncf': a.get('ncf'),
            'tipo_ncf': a.get('posiciones_fijas_ncf'),
            'posiciones_fijas_ncf': a.get('posiciones_fijas_ncf'),
            'fecha': a.get('fecha'), 'fecha_vence': a.get('fecha_vence'),
            'no_cia': a.get('no_cia'), 'punto': a.get('punto'),
            'detalle': a.get('detalle'), 'tipo_docu': 'CH', 'no_docu': a.get('no_docu'),
            'no_proveedor': a.get('no_bene'), 'no_reposicion': a.get('no_reposicion'),
            'tipo_gasto': a.get('tipo_gasto_dgii'),
            'valor_original': a.get('valor_original'), 'valor_total': a.get('valor_original'),
            'impuesto': a.get('impuesto'), 'itbis_facturado': a.get('impuesto'),
            'isc': a.get('isc'), 'otros_impuestos': a.get('otros_impuestos'),
            'propina': a.get('propina'),
            'itbis_retenido': 0, 'isr_retenido': 0, 'tipo_retencion': None,
            'valor_bienes': a.get('valor_bienes'), 'valor_servicio': a.get('valor_servicio'),
            'forma_pago': a.get('forma_pago'), 'modulo': 'ACC',
        })
    return rows


def _rows_acc_reposicion_606(no_cia: str, anio: int, mes: int, punto: str = ''):
    """Comprobante B13 (Gastos Menores) autoemitido por la empresa en cada
    reposicion de Caja Chica, para poder deducir los gastos SIN NCF de
    proveedor (compras informales) que componen esa reposicion. Vive en
    ACC.TACC_REPOSICION (no en TACC_DOCUMENTO) y tambien deben reportarse en
    el 606 (reportado 2026-08-12 por Pilar/JC: 'el numero de reposicion debe
    aparecer ahi' -- ej. reposicion 73 = B13-1301 $34,249.99, verificado que
    es la suma exacta de los 29 comprobantes sin NCF real de esa reposicion).
    El RNC/nombre del comprobante es el de la propia empresa (FAT.TFAT_CIAS),
    no un proveedor -- el B13 lo emite la empresa, no lo recibe.
    """
    conditions = [
        "r.no_cia=:1", "EXTRACT(YEAR FROM r.fecha)=:2", "EXTRACT(MONTH FROM r.fecha)=:3",
        "TRIM(r.ncf) IS NOT NULL", "r.anulada_por IS NULL",
    ]
    params = [no_cia, anio, mes]
    if punto:
        params.append(punto)
        conditions.append('r.punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT r.no_cia, r.punto, r.no_reposicion, r.ncf, r.posiciones_fijas_ncf, "
        "TO_CHAR(r.fecha,'YYYY-MM-DD') AS fecha, "
        "TO_CHAR(r.fecha_vence_ncf,'YYYY-MM-DD') AS fecha_vence, "
        "r.detalle, r.tipo_gasto_dgii, "
        "NVL(r.valor_ncf,0) valor_ncf, NVL(r.isc,0) isc, "
        "NVL(r.otros_impuestos,0) otros_impuestos, NVL(r.propina,0) propina, "
        "NVL(r.valor_bienes,0) valor_bienes, NVL(r.valor_servicio,0) valor_servicio, "
        "NVL(r.forma_pago,1) forma_pago "
        "FROM ACC.TACC_REPOSICION r WHERE " + where + " ORDER BY r.fecha, r.ncf"
    )
    raw = client.fetch_dicts(sql, params)
    if not raw:
        return []
    cia = client.fetch_one("SELECT rnc, descripcion FROM FAT.TFAT_CIAS WHERE no_cia=:1", [no_cia])
    rnc_empresa = (cia[0] if cia else '') or ''
    nombre_empresa = (cia[1] if cia else '') or ''
    rows = []
    for r in raw:
        rows.append({
            'rnc': rnc_empresa, 'rnc_proveedor': rnc_empresa,
            'nombre_proveedor': nombre_empresa,
            'tipo_transaccion': '',
            'ncf': r.get('ncf'), 'tipo_ncf': r.get('posiciones_fijas_ncf'),
            'posiciones_fijas_ncf': r.get('posiciones_fijas_ncf'),
            'fecha': r.get('fecha'), 'fecha_vence': r.get('fecha_vence'),
            'no_cia': r.get('no_cia'), 'punto': r.get('punto'),
            'detalle': r.get('detalle') or 'GASTOS MENORES CAJA CHICA',
            'tipo_docu': 'RC', 'no_docu': str(r.get('no_reposicion')),
            'no_proveedor': None, 'no_reposicion': r.get('no_reposicion'),
            'tipo_gasto': r.get('tipo_gasto_dgii'),
            'valor_original': r.get('valor_ncf'), 'valor_total': r.get('valor_ncf'),
            'impuesto': 0, 'itbis_facturado': 0,
            'isc': r.get('isc'), 'otros_impuestos': r.get('otros_impuestos'),
            'propina': r.get('propina'),
            'itbis_retenido': 0, 'isr_retenido': 0, 'tipo_retencion': None,
            'valor_bienes': r.get('valor_bienes'), 'valor_servicio': r.get('valor_servicio'),
            'forma_pago': r.get('forma_pago'), 'modulo': 'ACC',
        })
    return rows


def _ncf_modificado(no_cia: str, punto: str, tipo_docu: str, no_docu: str):
    """Para ND/NC con NCF propio (nota de debito/credito FISCAL emitida por
    el proveedor, no un ajuste interno), busca en TCXP_REFEDOCU el documento
    original que afecta y devuelve su NCF compuesto -- columna 'NCF
    Modificado' del 606 (reportado 2026-08-12 por Pilar/JC: 'donde tiene el
    NCF modificado sale su NCF y sale el NCF modificado... eso no esta
    apareciendo'). Toma la primera referencia si hay varias -- en los casos
    reales revisados una ND/NC con NCF propio siempre afecta un unico
    documento (a diferencia de las ND de aplicacion de pagos sin NCF propio,
    que sí pueden referenciar decenas de facturas y por eso quedan fuera del
    606 igualmente al no tener TRIM(ncf) IS NOT NULL).
    """
    from .fat_repo import _compose_ncf_dgi
    row = client.fetch_one(
        "SELECT d2.posiciones_fijas_ncf, d2.ncf FROM CXP.TCXP_REFEDOCU r "
        "JOIN CXP.TCXP_DOCUMENTO d2 ON d2.no_cia=r.no_cia AND d2.punto=r.punto "
        "AND d2.tipo_docu=r.tipo_refe AND d2.no_docu=r.no_refe "
        "WHERE r.no_cia=:1 AND r.punto=:2 AND r.tipo_docu=:3 AND r.no_docu=:4 "
        "AND d2.ncf IS NOT NULL "
        "ORDER BY r.tipo_refe, r.no_refe FETCH FIRST 1 ROW ONLY",
        [no_cia, punto, tipo_docu, no_docu])
    if not row:
        return ''
    return _compose_ncf_dgi(row[0], row[1]) or ''


def rep_606(no_cia: str, anio: int, mes: int, punto: str = ''):
    """Reporte 606 - ITBIS en compras locales (Formato DGII).

    Solo incluye documentos con NCF registrado: el usuario reporto que estaba
    devolviendo tambien gastos sin NCF, y para el 606 solo aplican los que
    tienen comprobante fiscal. Se excluyen B02/E32 (Factura de Consumo): ese
    tipo de NCF lo EMITE la empresa al vender a un consumidor final, no algo
    que deba recibir como comprobante de compra -- si aparece uno es un
    error de captura del proveedor/informal. Incluye tambien los
    comprobantes de Caja Chica (ACC) con NCF via _rows_acc_606().

    tipo_docu IN ('FP','FT','NC','ND') en vez del viejo filtro tipo_movi='C':
    ese filtro excluia TODAS las Notas de Debito (tipo_movi='D') aunque
    tuvieran NCF fiscal real del proveedor -- confirmado con datos reales de
    julio/2026 (3 ND con NCF E34/B04 que nunca aparecian). AC/AD/BD/BI/SO
    son ajustes/balances/solicitudes internas, no comprobantes de compra.
    """
    conditions = [
        "d.no_cia=:1",
        "EXTRACT(YEAR FROM d.fecha)=:2",
        "EXTRACT(MONTH FROM d.fecha)=:3",
        "d.tipo_docu IN ('FP','FT','NC','ND')",
        "TRIM(d.ncf) IS NOT NULL",
        "NVL(UPPER(d.posiciones_fijas_ncf),'') NOT IN ('B02','E32')",
        # Un documento reversado (status='R') no representa una compra real
        # ante la DGII -- el mismo patron usado para NCF duplicados (ver
        # _check_ncf_duplicate) aplica aqui: no debe listarse ni exportarse.
        "NVL(d.status,'A') <> 'R'",
    ]
    params = [no_cia, anio, mes]
    if punto:
        params.append(punto)
        conditions.append('d.punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT d.rnc AS rnc_proveedor, "
        "p.nombre AS nombre_proveedor, "
        "d.tipo_transaccion, "
        "d.ncf, d.posiciones_fijas_ncf AS tipo_ncf, "
        "TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha, "
        "TO_CHAR(d.fecha_vence,'YYYY-MM-DD') AS fecha_vence, "
        "d.no_cia, d.punto, d.detalle, d.tipo_docu, d.no_docu, d.no_proveedor, d.tipo_gasto, "
        # Campos crudos para _base_606 (Monto Facturado = base sin impuestos).
        "NVL(d.valor_original,0) valor_original, NVL(d.impuesto,0) impuesto, "
        "NVL(d.isc,0) isc, NVL(d.otros_impuestos,0) otros_impuestos, "
        "NVL(d.propina,0) propina, "
        # Alias de presentacion (compatibilidad con frontend/print-data).
        "NVL(d.valor_original,0) valor_total, "
        "NVL(d.impuesto,0) itbis_facturado, "
        "NVL(d.itbis_retenido,0) itbis_retenido, "
        "NVL(d.isr_retenido,0) isr_retenido, "
        "d.tipo_retencion, "
        "NVL(d.valor_bienes,0) valor_bienes, "
        "NVL(d.valor_servicio,0) valor_servicio, "
        "NVL(d.forma_pago,1) forma_pago "
        "FROM CXP.TCXP_DOCUMENTO d "
        "LEFT JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor=d.no_proveedor "
        "WHERE " + where + " ORDER BY d.fecha, d.ncf"
    )
    rows = client.fetch_dicts(sql, params)
    for r in rows:
        r['modulo'] = 'CXP'
    rows += _rows_acc_606(no_cia, anio, mes, punto)
    rows += _rows_acc_reposicion_606(no_cia, anio, mes, punto)
    rows.sort(key=lambda r: (r.get('fecha') or '', str(r.get('ncf') or '')))
    for r in rows:
        base = _base_606(r)
        r['monto_facturado'] = base
        # Servicios vs Bienes por valor_servicio (>0 => servicio); regla del
        # legado verificada contra el 606 real (no depende de tipo_gasto).
        es_serv = float(r.get('valor_servicio') or 0) > 0
        r['monto_servicios'] = base if es_serv else 0.0
        r['monto_bienes'] = 0.0 if es_serv else base
        r['tipo_gasto_label'] = _tipo_gasto_label(r.get('tipo_gasto'))
        if r.get('tipo_docu') in ('ND', 'NC') and r.get('ncf'):
            r['ncf_modificado'] = _ncf_modificado(
                r.get('no_cia'), r.get('punto'), r.get('tipo_docu'), r.get('no_docu'))
    total_monto = sum((r.get('monto_facturado') or 0) for r in rows)
    total_itbis = sum((r.get('itbis_facturado') or 0) for r in rows)
    total_servicios = sum((r.get('monto_servicios') or 0) for r in rows)
    total_bienes = sum((r.get('monto_bienes') or 0) for r in rows)
    return {'items': rows, 'count': len(rows),
            'total_monto': total_monto, 'total_itbis': total_itbis,
            'total_servicios': total_servicios, 'total_bienes': total_bienes}


def archivo_dgii_606(no_cia: str, anio: int, mes: int, punto: str = '') -> tuple[str, int]:
    """Genera el contenido del archivo 606 (formato DGII, pipe-delimited) tal
    como lo produce el legado (ver C:\\archivo_ncf\\archivo_606_*.txt).

    Limitacion conocida (no resuelta, revisar con un contador antes de
    enviar a DGII): la columna Monto Facturado en Servicios vs en Bienes
    (campos 8/9) no tiene una fuente confiable en el esquema actual --
    TCXP_DOCUMENTO.valor_bienes/valor_servicio existen pero
    entrada_documento() no los llena para documentos nuevos. Se usa una
    regla aproximada: tipo_gasto='03' (Arrendamientos) va a Servicios,
    todo lo demas va a Bienes -- coincide con los ejemplos reales
    revisados pero no esta garantizado para todos los casos.
    """
    from .fat_repo import _compose_ncf_dgi, _tipo_id_de_rnc
    conditions = [
        "d.no_cia=:1", "EXTRACT(YEAR FROM d.fecha)=:2", "EXTRACT(MONTH FROM d.fecha)=:3",
        # ver docstring de rep_606(): tipo_docu IN (...) en vez de tipo_movi='C',
        # que excluia las Notas de Debito (D) aunque tuvieran NCF fiscal real.
        "d.tipo_docu IN ('FP','FT','NC','ND')", "TRIM(d.ncf) IS NOT NULL",
        # B02/E32 (Factura de Consumo) no debe aparecer como comprobante de
        # compra -- la empresa lo emite al vender, no lo recibe al comprar.
        "NVL(UPPER(d.posiciones_fijas_ncf),'') NOT IN ('B02','E32')",
        # Reversado (status='R') = no es una compra real, no va al archivo DGII.
        "NVL(d.status,'A') <> 'R'",
    ]
    params: list = [no_cia, anio, mes]
    if punto:
        params.append(punto)
        conditions.append('d.punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    rows = client.fetch_dicts(
        "SELECT d.rnc, d.ncf, d.posiciones_fijas_ncf, d.tipo_gasto, "
        "d.no_cia, d.punto, d.tipo_docu, d.no_docu, "
        "TO_CHAR(d.fecha,'YYYYMMDD') fecha, "
        "NVL(d.valor_original,0) valor_original, NVL(d.impuesto,0) impuesto, "
        "NVL(d.itbis_retenido,0) itbis_retenido, NVL(d.isr_retenido,0) isr_retenido, "
        "d.tipo_retencion, NVL(d.isc,0) isc, NVL(d.otros_impuestos,0) otros_impuestos, "
        "NVL(d.propina,0) propina, NVL(d.forma_pago,4) forma_pago, "
        "NVL(d.valor_servicio,0) valor_servicio "
        "FROM CXP.TCXP_DOCUMENTO d WHERE " + where + " ORDER BY d.fecha, d.ncf",
        params)
    rows += _rows_acc_606(no_cia, anio, mes, punto)
    rows += _rows_acc_reposicion_606(no_cia, anio, mes, punto)
    rows.sort(key=lambda r: (r.get('fecha') or '', str(r.get('ncf') or '')))

    cia = client.fetch_one("SELECT rnc FROM FAT.TFAT_CIAS WHERE no_cia=:1", [no_cia])
    rnc_empresa = (cia[0] if cia else '') or ''

    lineas = []
    for r in rows:
        ncf_dgi = _compose_ncf_dgi(r['posiciones_fijas_ncf'], r['ncf'])
        tipo_gasto = f"{int(r['tipo_gasto']):02d}" if str(r['tipo_gasto'] or '').isdigit() else '02'
        # Servicios vs Bienes: el legado clasifica por valor_servicio (>0 =>
        # servicio), NO por tipo_gasto. Verificado contra el 606 real de julio
        # (RMM honorarios y arrendamientos salen en Servicios aunque su
        # tipo_gasto sea 02/03; el resto en Bienes). Ver _base_606().
        es_servicio = float(r['valor_servicio'] or 0) > 0
        impuesto = float(r['impuesto'])
        base = _base_606(r)
        monto_servicios = base if es_servicio else 0.0
        monto_bienes = 0.0 if es_servicio else base
        ncf_modificado = ''
        if r.get('tipo_docu') in ('ND', 'NC') and r.get('ncf'):
            ncf_modificado = _ncf_modificado(
                r.get('no_cia'), r.get('punto'), r.get('tipo_docu'), r.get('no_docu'))
        # Filas de CXP ya vienen en YYYYMMDD; las de ACC (_rows_acc_606/
        # _rows_acc_reposicion_606) en YYYY-MM-DD -- quitar guiones normaliza
        # ambas sin reformatear (evita fechas con guion coladas en el TXT).
        fecha_ymd = (r.get('fecha') or '').replace('-', '')
        campos = [
            r['rnc'] or '', _tipo_id_de_rnc(r['rnc']), tipo_gasto, ncf_dgi, ncf_modificado,
            fecha_ymd, fecha_ymd,
            f"{monto_servicios:.2f}", f"{monto_bienes:.2f}", f"{base:.2f}",
            f"{impuesto:.2f}", f"{float(r['itbis_retenido']):.2f}",
            '0', '0.00', f"{float(r['impuesto']):.2f}", '0',
            str(int(r['tipo_retencion'])) if r['tipo_retencion'] else '',
            f"{float(r['isr_retenido']):.2f}", '0',
            f"{float(r['isc']):.2f}", f"{float(r['otros_impuestos']):.2f}",
            f"{float(r['propina']):.2f}", f"{int(r['forma_pago']):02d}",
        ]
        lineas.append('|'.join(campos))

    contenido = f"606|{rnc_empresa}|{anio}{mes:02d}|{len(lineas)}\n" + '\n'.join(lineas) + ('\n' if lineas else '')
    return contenido, len(lineas)


# Encabezados del Excel 606 — orden y textos identicos al reporte del legado
# (REPORTE_606_MMYYYY.ods) para que "meta los mismos campos".
_EXCEL_606_HEADERS = [
    'RNC/CED', 'TIPO ID', 'TIPO GTO', 'NCF', 'NCF MODIFICADO',
    'FECHA_COMPROBANTE', 'DIA (FECHA COMPROBANTE)', 'MODULO', 'DOCUMENTO',
    'NO. REPOS.', 'NOMBRE', 'DETALLE', 'FECHA PAGO', 'DIA (FECHA PAGO)',
    'MONTO SERVICIOS', 'MONTO BIENES', 'MONTO FACTURADO', 'ITBIS FACTURADO',
    'ITBIS RETENIDO', 'ITBIS PROPORCIONALIDAD', 'ITBIS LLEVADO COSTO',
    'ITBIS POR ADELANTAR', 'ITBIS PERCIB. COMPRAS', 'TIPO RETENCION',
    'ISR RETENIDO', 'ISR PERCIB. COMPRAS', 'ISC', 'OTROS IMPUESTOS/TASAS',
    'PROPINA', 'FORMA PAGO', 'FECHA PROVEEDOR',
]


def rep_606_excel(no_cia: str, anio: int, mes: int, punto: str = '') -> bytes:
    """Genera el Excel (.xlsx) del 606 con las mismas columnas del reporte del
    legado. Reusa rep_606() (mismos montos/servicios/bienes) para no duplicar
    la logica de la base ni de la clasificacion servicio/bien."""
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from .fat_repo import _compose_ncf_dgi, _tipo_id_de_rnc

    data = rep_606(no_cia, anio, mes, punto)
    cia = client.fetch_one("SELECT descripcion FROM FAT.TFAT_CIAS WHERE no_cia=:1", [no_cia])
    razon = (cia[0] if cia else '') or ''

    def _ymd(s):  # 'YYYY-MM-DD' -> ('YYYYMMDD', dia, 'D/M/YYYY')
        try:
            d = _dt.strptime(str(s)[:10], '%Y-%m-%d')
            return d.strftime('%Y%m%d'), d.day, f"{d.day}/{d.month}/{d.year}"
        except Exception:
            return str(s or ''), '', ''

    wb = Workbook()
    ws = wb.active
    ws.title = '606'
    ncols = len(_EXCEL_606_HEADERS)
    ws.append([razon]); ws['A1'].font = Font(bold=True, size=12)
    ws.append([f"Reporte de Documentos con NCF. Formato 606 Desde 01/{mes:02d}/{anio} "
               f"Hasta 31/{mes:02d}/{anio}"])
    ws.append([])
    ws.append(list(_EXCEL_606_HEADERS))
    for c in ws[4]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    money_cols = set(range(15, 30))  # 1-based columnas de montos → formato numerico
    for r in data['items']:
        ymd, dia, dprov = _ymd(r.get('fecha'))
        ncf_dgi = (_compose_ncf_dgi(r.get('tipo_ncf'), r.get('ncf'))
                   or (str(r.get('ncf')) if r.get('ncf') else ''))
        documento = f"{r.get('no_cia')}-{r.get('punto')}-{r.get('tipo_docu')}-{r.get('no_docu')}"
        itbis_fact = float(r.get('itbis_facturado') or 0)
        tr = r.get('tipo_retencion')
        tipo_ret = str(int(tr)) if tr not in (None, '', 0, '0') else ''
        fila = [
            r.get('rnc_proveedor') or '', _tipo_id_de_rnc(r.get('rnc_proveedor')),
            r.get('tipo_gasto_label') or '', ncf_dgi, r.get('ncf_modificado') or '',
            ymd, dia, r.get('modulo') or 'CXP', documento, r.get('no_reposicion') or '',
            r.get('nombre_proveedor') or '', r.get('detalle') or '',
            ymd, dia,
            float(r.get('monto_servicios') or 0), float(r.get('monto_bienes') or 0),
            float(r.get('monto_facturado') or 0), itbis_fact,
            float(r.get('itbis_retenido') or 0), 0, 0, itbis_fact, 0, tipo_ret,
            float(r.get('isr_retenido') or 0), 0,
            float(r.get('isc') or 0), float(r.get('otros_impuestos') or 0),
            float(r.get('propina') or 0),
            f"{int(r.get('forma_pago') or 0):02d}", dprov,
        ]
        ws.append(fila)
        for ci in money_cols:
            ws.cell(row=ws.max_row, column=ci).number_format = '#,##0.00'

    # Fila de totales
    tot_row = ['' for _ in range(ncols)]
    tot_row[10] = 'TOTALES'
    tot_row[14] = float(data.get('total_servicios') or 0)
    tot_row[15] = float(data.get('total_bienes') or 0)
    tot_row[16] = float(data.get('total_monto') or 0)
    tot_row[17] = float(data.get('total_itbis') or 0)
    ws.append(tot_row)
    for ci in (11, 15, 16, 17, 18):
        c = ws.cell(row=ws.max_row, column=ci)
        c.font = Font(bold=True)
        if ci >= 15:
            c.number_format = '#,##0.00'

    widths = [13, 7, 42, 16, 14, 14, 10, 8, 20, 10, 30, 30, 12, 10] + [15] * 17
    for i, w in enumerate(widths[:ncols], start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rep_607(no_cia: str, anio: int, mes: int, punto: str = ''):
    """Reporte 607 - Retenciones del ISR (Formato DGII)."""
    conditions = [
        "d.no_cia=:1",
        "EXTRACT(YEAR FROM d.fecha)=:2",
        "EXTRACT(MONTH FROM d.fecha)=:3",
        # La retencion se registra en la factura del proveedor (tipo_movi='C'),
        # no en el pago: verificado 2026-07-02 — todos los docs con
        # isr_retenido>0 en produccion son 'C'.
        "d.tipo_movi='C'",
        "NVL(d.isr_retenido,0)>0",
    ]
    params = [no_cia, anio, mes]
    if punto:
        params.append(punto)
        conditions.append('d.punto=:' + str(len(params)))
    where = ' AND '.join(conditions)
    sql = (
        "SELECT d.rnc AS rnc_proveedor, "
        "p.nombre AS nombre_proveedor, "
        "d.tipo_transaccion, d.ncf, "
        "TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha, "
        "NVL(d.valor_original,0) monto_pago, "
        "NVL(d.isr_retenido,0) isr_retenido, "
        "NVL(d.itbis_retenido,0) itbis_retenido, "
        "NVL(d.tipo_retencion,1) tipo_retencion, "
        "d.no_proveedor, d.tipo_docu, d.no_docu "
        "FROM CXP.TCXP_DOCUMENTO d "
        "LEFT JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor=d.no_proveedor "
        "WHERE " + where + " ORDER BY d.fecha, d.no_docu"
    )
    rows = client.fetch_dicts(sql, params)
    total_isr   = sum((r.get('isr_retenido')   or 0) for r in rows)
    total_itbis = sum((r.get('itbis_retenido') or 0) for r in rows)
    return {'items': rows, 'count': len(rows),
            'total_isr': total_isr, 'total_itbis': total_itbis}


# ─── PROCESOS ESCRITURA ─────────────────────────────────────────────────────


def get_siguiente_no_docu(no_cia, punto, tipo_docu):
    """Preview del siguiente numero de doc (sin commit, solo lectura).

    SEMANTICA LEGACY: TCXP_SECUENCIA.ult_docu = 'siguiente a usar' (NO
    'ultimo usado'). El form legacy lee ult_docu, lo usa como no_docu del
    documento nuevo, y DESPUES incrementa. Por eso NO sumamos 1 aqui.
    """
    rows = client.fetch_dicts(
        "SELECT NVL(ult_docu,0) AS siguiente FROM CXP.TCXP_SECUENCIA "
        "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3",
        [no_cia, punto, tipo_docu])
    n = int(rows[0]['siguiente']) if rows else 1
    # Si el row no existe todavia, el primer no_docu sera 1.
    if n == 0:
        n = 1
    return str(n).zfill(7)


def _next_no_docu(cur, no_cia, punto, tipo_docu):
    """Reserva el siguiente no_docu y avanza TCXP_SECUENCIA (FOR UPDATE).

    Semantica legacy: ult_docu = numero a usar AHORA. Tras la INSERT,
    ult_docu pasa a ser el siguiente disponible (ult_docu + 1).
    """
    rows = cur.execute(
        "SELECT ult_docu FROM CXP.TCXP_SECUENCIA "
        "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 FOR UPDATE",
        [no_cia, punto, tipo_docu]).fetchone()
    if rows:
        a_usar = rows[0] or 1
        cur.execute(
            "UPDATE CXP.TCXP_SECUENCIA SET ult_docu=:1 "
            "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4",
            [a_usar + 1, no_cia, punto, tipo_docu])
    else:
        # Hay no_docu no numericos en datos migrados — filtrar antes de TO_NUMBER
        row_max = cur.execute(
            "SELECT NVL(MAX(TO_NUMBER(no_docu)),0) FROM CXP.TCXP_DOCUMENTO "
            "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 "
            "  AND REGEXP_LIKE(no_docu, '^[0-9]+$')",
            [no_cia, punto, tipo_docu]).fetchone()
        a_usar = (row_max[0] if row_max else 0) + 1
        cur.execute(
            "INSERT INTO CXP.TCXP_SECUENCIA(no_cia,punto,tipo_docu,ult_docu) "
            "VALUES(:1,:2,:3,:4)",
            [no_cia, punto, tipo_docu, a_usar + 1])
    return str(int(a_usar)).zfill(7)


def _check_ncf_duplicate(cur, no_cia, no_proveedor, ncf_num, pos_ncf,
                          exclude=None):
    """Rechaza registrar dos veces el mismo NCF (prefijo+numero) para el
    mismo proveedor dentro de la misma empresa. Salta si el NCF esta vacio
    (ajustes/cheques sin comprobante fiscal). Ignora documentos reversados
    (status='R') porque su NCF puede re-registrarse en el correcto.

    exclude=(tipo_docu, no_docu) evita que un UPDATE colisione consigo mismo.
    """
    if ncf_num is None or not pos_ncf:
        return
    sql = (
        "SELECT tipo_docu, no_docu FROM CXP.TCXP_DOCUMENTO "
        "WHERE no_cia=:1 AND no_proveedor=:2 AND ncf=:3 "
        "  AND UPPER(posiciones_fijas_ncf)=:4 "
        "  AND NVL(status,'A') <> 'R'"
    )
    params = [no_cia, str(no_proveedor), ncf_num, str(pos_ncf).upper()]
    if exclude:
        sql += " AND NOT (tipo_docu=:5 AND no_docu=:6)"
        params += [exclude[0], exclude[1]]
    row = cur.execute(sql, params).fetchone()
    if row:
        _pos_up = str(pos_ncf).upper()
        _width  = 10 if _pos_up.startswith('E') else 8
        raise ValueError(
            "El NCF {0}{1} ya esta registrado para este proveedor "
            "(documento {2} {3}). No se permite duplicar NCF en CxP.".format(
                _pos_up, str(ncf_num).zfill(_width),
                row[0], row[1],
            )
        )


def _ensure_bproveedor(cur, no_cia, punto, no_proveedor, usuario='API'):
    """INSERT CXP.TCXP_BPROVEEDOR si falta — mismo patron que
    inv_repo._insert_eproducto para TINV_EPRODUCTO. TCXP_DOCUMENTO tiene
    FK_TCXP_DOCUMENTO_BPROVE contra (no_cia,punto,no_proveedor) de esta
    tabla, pero save_proveedor() solo inserta en TCXP_DPROVEEDOR (que es
    global, sin no_cia/punto) -- cualquier proveedor creado desde el clon
    revienta con ORA-02291 al grabar su primer documento en una cia/punto
    donde nunca tuvo balance. Ver reporte de soporte de ACLASE 2026-08-13.
    Idempotente."""
    exists = cur.execute(
        "SELECT 1 FROM CXP.TCXP_BPROVEEDOR "
        "WHERE no_cia=:1 AND punto=:2 AND no_proveedor=:3",
        [no_cia, punto, no_proveedor]).fetchone()
    if exists:
        return
    cur.execute(
        "INSERT INTO CXP.TCXP_BPROVEEDOR("
        "no_cia,punto,no_proveedor,tipo_proveedor,activo,"
        "saldo_inicial,debitos,creditos,fecha,usuario"
        ") VALUES(:1,:2,:3,'01','S',0,0,0,SYSDATE,:4)",
        [no_cia, punto, no_proveedor, usuario])


def entrada_documento(d, _skip_periodo_gate: bool = False):
    """
    Crea/actualiza un documento CxP (cabecera TCXP_DOCUMENTO + lineas TCXP_DCDOCU).
    Tipos habituales: AC, AD, BD, FP. REVERSIBLE via reversar_documento.

    Candado de periodo (solo al CREAR, no al editar; se salta con
    _skip_periodo_gate=True durante la materializacion de la cola):
      - CERRADO (fecha de un periodo anterior al de proceso): se BLOQUEA con
        ValueError -- back-datear a un mes ya cerrado desalinearia el 606 ya
        presentado a la DGII.
      - FUTURO (fecha de un periodo posterior al de proceso): se ENCOLA en
        TCXP_DOCUMENTO_COLA y se levanta PeriodoFuturoEncolado. La entrada
        de INV/FAT no se rompe; el FP se materializa al abrir ese mes.
      - ABIERTO (mismo periodo): flujo normal de insercion.
    """
    no_cia        = d["no_cia"]
    punto         = d.get("punto", "01")
    tipo_docu     = d["tipo_docu"]
    no_proveedor  = str(d["no_proveedor"])
    # Normalizar fechas ANTES de armar el SQL: un año de 5 dígitos (26810)
    # provenía del <input type="date"> y reventaba TO_DATE con ORA-01861.
    fecha         = _norm_fecha(d.get("fecha", ""), campo="fecha", requerido=True)
    fecha_vence   = _norm_fecha(d.get("fecha_vence") or fecha, campo="fecha de vencimiento") or fecha
    valor         = float(d.get("valor_original") or d.get("valor") or 0)
    # TIPO_MOVI ('D'=debito / 'C'=credito) es una propiedad fija del tipo de
    # documento (TCXP_TDOCU.TIPO_MOVI, ej. ND/AD/BD='D', FP/NC/AC='C'), NO
    # una eleccion del usuario. El frontend de Entrada de Documentos nunca
    # envia "tipo_movi", asi que confiar en d.get("tipo_movi","C") hacia
    # SIEMPRE 'C' para cualquier tipo -- invisible en FP (su default real
    # tambien es 'C') pero silenciosamente mal para ND/AD/BD ('D'), dejando
    # esos documentos con signo de saldo invertido (ver reporte de soporte
    # de MPILAR: Nota de Debito no aparecia como saldo a favor al aplicar).
    _tdocu_rows = client.fetch_dicts(
        "SELECT tipo_movi FROM CXP.TCXP_TDOCU WHERE tipo_docu=:1", [tipo_docu])
    if _tdocu_rows:
        tipo_movi = (_tdocu_rows[0]["tipo_movi"] or "C").upper()
    else:
        _tm_raw   = d.get("tipo_movi", "C")
        tipo_movi = "D" if str(_tm_raw).upper() in ("D", "DEBITO", "DEBIT") else "C"
    # TIPO_TRANSACCION is VARCHAR2(1): 'K'=compra local, 'C'=compra exterior
    # Accept legacy '01'/'02' or short 'K'/'C' forms
    _tt = str(d.get("tipo_transaccion") or "K").strip()
    tipo_transacc = _tt[0] if len(_tt) == 1 else ("C" if _tt in ("02", "C", "c") else "K")
    detalle       = d.get("detalle", "")
    no_docu       = (d.get("no_docu") or "").strip()

    # Candado de periodo: solo al CREAR (no_docu vacio) y salvo materializacion.
    if not no_docu and not _skip_periodo_gate:
        _estado_per, _per_proc = _periodo_estado(no_cia, punto, fecha)
        if _estado_per == 'CERRADO':
            raise ValueError(
                'La fecha del documento ({0}) pertenece a un periodo contable '
                'ya CERRADO en CxP; el periodo en curso es {1}. No se puede '
                'registrar con esa fecha.'.format(fecha[:7], _per_proc))
        if _estado_per == 'FUTURO':
            _origen = 'INV' if d.get('_cola_link') else d.get('_cola_origen', 'MANUAL')
            _cid = _encolar_documento({**d, 'fecha': fecha, 'fecha_vence': fecha_vence,
                                       'no_cia': no_cia, 'punto': punto}, origen=_origen)
            raise PeriodoFuturoEncolado(_cid, fecha[:7], _per_proc)

    with client.cursor() as cur:
        # Oracle constraint: debito=credito always; saldo<=0 for D, >=0 for C
        saldo_inicial = -valor if tipo_movi == "D" else valor
        _ensure_bproveedor(cur, no_cia, punto, no_proveedor, d.get("usuario", "API"))

        if no_docu:
            # Editar un documento ya existente (usado por el boton "Editar"
            # de Consulta de Documentos, que manda al usuario de vuelta a
            # esta pantalla con todo precargado). Dos candados antes de
            # permitirlo:
            #  1) Solo el periodo contable en curso -- editar un periodo ya
            #     cerrado desalinearia el 606 ya presentado (mismo criterio
            #     que corregir_datos_dgii).
            #  2) El documento no puede tener ya un pago/aplicacion parcial
            #     (saldo != valor_original original) -- sobreescribir
            #     valor_original/saldo aqui borraria el rastro de esa
            #     aplicacion. Hay que revertirla primero (Reversar/Aplicacion
            #     de Movimientos) en vez de editar por encima.
            _actual = client.fetch_dicts(
                "SELECT valor_original, saldo, "
                "TO_CHAR(fecha,'YYYY-MM') AS periodo_docu "
                "FROM CXP.TCXP_DOCUMENTO "
                "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
                [no_cia, punto, tipo_docu, no_docu])
            if not _actual:
                raise ValueError(f"Documento {tipo_docu}-{no_docu} no encontrado")
            _valor_orig_actual = float(_actual[0]['valor_original'] or 0)
            _saldo_actual = float(_actual[0]['saldo'] or 0)
            if round(abs(_saldo_actual), 2) != round(_valor_orig_actual, 2):
                raise ValueError(
                    "Este documento ya tiene pagos o aplicaciones parciales "
                    "(saldo distinto del valor original) -- revierta esos "
                    "movimientos antes de editarlo directamente."
                )
            _punto_rows = client.fetch_dicts(
                "SELECT ano_proceso, mes_proceso FROM CXP.TCXP_PUNTO "
                "WHERE no_cia=:1 AND punto=:2", [no_cia, punto])
            if _punto_rows:
                _periodo_actual = '{:04d}-{:02d}'.format(
                    int(_punto_rows[0]['ano_proceso']), int(_punto_rows[0]['mes_proceso']))
                # Bloquear solo si el documento es de un periodo YA CERRADO
                # (anterior al periodo de proceso). Comparar con != tambien
                # bloqueaba documentos del mes calendario actual cuando el
                # punto todavia no habia avanzado su mes_proceso (cierre
                # pendiente de ejecutar) -- eso no es un periodo cerrado.
                if _actual[0]['periodo_docu'] < _periodo_actual:
                    raise ValueError(
                        'Este documento pertenece a un periodo contable ya '
                        'cerrado ({1}); el periodo en curso es {0}.'.format(
                            _periodo_actual, _actual[0]['periodo_docu']))

            _ncf_raw_u = str(d.get("ncf") or '').strip()
            _ncf_num_u = int(_ncf_raw_u) if _ncf_raw_u.isdigit() else None
            _pos_ncf_u = (
                str(d.get("posiciones_fijas_ncf") or d.get("tipo_ncf") or '')
                .strip().upper() or None
            )
            _check_ncf_duplicate(cur, no_cia, no_proveedor, _ncf_num_u, _pos_ncf_u,
                                 exclude=(tipo_docu, no_docu))
            cur.execute(
                "UPDATE CXP.TCXP_DOCUMENTO SET "
                "no_proveedor=:1, fecha=TO_DATE(:2,'YYYY-MM-DD'), "
                "fecha_vence=TO_DATE(:3,'YYYY-MM-DD'), "
                "valor_original=:4, debito=:5, credito=:6, saldo=:7, "
                "tipo_movi=:8, tipo_transaccion=:9, detalle=:10, "
                "ncf=:11, rnc=:12, impuesto=:13, "
                "itbis_retenido=:14, isr_retenido=:15, forma_pago=:16, "
                "valor_bienes=:17, valor_servicio=:18 "
                "WHERE no_cia=:19 AND punto=:20 AND tipo_docu=:21 AND no_docu=:22",
                [
                    no_proveedor, fecha, fecha_vence,
                    valor,          # :4 valor_original
                    valor,          # :5 debito (= valor_original always)
                    valor,          # :6 credito (= valor_original always)
                    saldo_inicial,  # :7 saldo
                    tipo_movi, tipo_transacc, detalle,
                    d.get("ncf", ""), d.get("rnc", ""),
                    float(d.get("impuesto") or 0),
                    float(d.get("itbis_retenido") or 0),
                    float(d.get("isr_retenido") or 0),
                    int(d.get("forma_pago") or 1),
                    float(d.get("valor_bienes") or 0) or None,
                    float(d.get("valor_servicio") or 0) or None,
                    no_cia, punto, tipo_docu, no_docu,
                ])
        else:
            # NCF: el usuario teclea solo los digitos (o el sistema autoasigna
            # desde TCNT_NCF.PROX_NCF). Lo guardamos como NUMBER y el prefijo
            # (B11/B01/etc) va en POSICIONES_FIJAS_NCF.
            _ncf_raw = str(d.get("ncf") or '').strip()
            _ncf_num = int(_ncf_raw) if _ncf_raw.isdigit() else None
            _pos_ncf = (
                d.get("posiciones_fijas_ncf")
                or d.get("tipo_ncf")
                or ''
            ).strip().upper() or None
            # Validar duplicado ANTES de reservar no_docu para no dejar huecos
            # en TCXP_SECUENCIA cuando el operador reintenta con el mismo NCF.
            _check_ncf_duplicate(cur, no_cia, no_proveedor, _ncf_num, _pos_ncf)
            no_docu = _next_no_docu(cur, no_cia, punto, tipo_docu)
            # Campos DGI adicionales (opcionales, NULL si no se envian).
            _isc        = float(d.get("isc") or 0) or None
            _otros_imp  = float(d.get("otros_impuestos") or 0) or None
            _propina    = float(d.get("propina") or 0) or None
            _tipo_gasto = (str(d.get("tipo_gasto") or '').strip() or None)
            _tipo_ret   = d.get("tipo_retencion")
            _tipo_ret   = int(_tipo_ret) if _tipo_ret not in (None, '', 0) else None
            _forma_pago = d.get("forma_pago")
            _forma_pago = int(_forma_pago) if _forma_pago not in (None, '') else None
            _valor_bienes   = float(d.get("valor_bienes") or 0) or None
            _valor_servicio = float(d.get("valor_servicio") or 0) or None
            cur.execute(
                "INSERT INTO CXP.TCXP_DOCUMENTO("
                "no_cia,punto,tipo_docu,no_docu,no_proveedor,tipo_movi,tipo_transaccion,"
                "fecha,fecha_vence,status,valor_original,debito,credito,saldo,"
                "impuesto,itbis_retenido,isr_retenido,rnc,ncf,posiciones_fijas_ncf,detalle,"
                "isc,otros_impuestos,propina,tipo_gasto,tipo_retencion,forma_pago,"
                "valor_bienes,valor_servicio,"
                "st_generado_cnt,pago_bloqueado,estado_encf,usuario,fecha_sysdate"
                ") VALUES("
                ":1,:2,:3,:4,:5,:6,:7,"
                "TO_DATE(:8,'YYYY-MM-DD'),TO_DATE(:9,'YYYY-MM-DD'),'A',:10,:11,:12,:13,"
                ":14,:15,:16,:17,:18,:19,:20,"
                ":21,:22,:23,:24,:25,:26,"
                ":27,:28,"
                "'N','N',0,:29,SYSDATE"
                ")",
                [
                    no_cia, punto, tipo_docu, no_docu, no_proveedor,
                    tipo_movi, tipo_transacc,
                    fecha, fecha_vence,
                    valor,          # :10 valor_original
                    valor,          # :11 debito (= valor_original always)
                    valor,          # :12 credito (= valor_original always)
                    saldo_inicial,  # :13 saldo
                    float(d.get("impuesto") or 0),
                    float(d.get("itbis_retenido") or 0),
                    float(d.get("isr_retenido") or 0),
                    d.get("rnc", ""), _ncf_num, _pos_ncf, detalle,
                    _isc, _otros_imp, _propina,
                    _tipo_gasto, _tipo_ret, _forma_pago,
                    _valor_bienes, _valor_servicio,
                    d.get("usuario", "API"),
                ])

        # El legado nunca deja grabar un documento sin su partida doble
        # (ver Fcxp201: no permite COMMIT si Debe<>Haber). El API se estaba
        # usando sin este candado -- tanto la UI antigua como llamadas
        # directas (pruebas de integracion INV->CXP, etc.) podian grabar
        # TCXP_DOCUMENTO sin ninguna linea en TCXP_DCDOCU, dejando el
        # documento sin movimiento contable ni rastro en 606/mayor (ver
        # FP-0008558..0008594 y FP-0008619). Se valida aqui, no solo en el
        # frontend, para cubrir cualquier caller presente o futuro.
        _lineas_in = d.get("lineas", [])
        if not _lineas_in:
            raise ValueError(
                "El documento debe incluir la distribucion contable "
                "(lineas de Debito/Credito) -- no se puede grabar sin ella.")
        _tot_deb = sum(float(l.get("monto") or 0) for l in _lineas_in
                       if (l.get("tipo_movi") or "D").upper() == "D")
        _tot_cre = sum(float(l.get("monto") or 0) for l in _lineas_in
                       if (l.get("tipo_movi") or "D").upper() == "C")
        if round(abs(_tot_deb - _tot_cre), 2) > 0.01:
            raise ValueError(
                "La distribucion contable no cuadra: Debito {:.2f} vs "
                "Credito {:.2f}.".format(_tot_deb, _tot_cre))

        cur.execute(
            "DELETE FROM CXP.TCXP_DCDOCU "
            "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
            [no_cia, punto, tipo_docu, no_docu])

        no_prov_num = int(no_proveedor) if no_proveedor.isdigit() else 0
        for linea in d.get("lineas", []):
            tm    = linea.get("tipo_movi", "D")
            monto = float(linea.get("monto") or 0)
            cur.execute(
                "INSERT INTO CXP.TCXP_DCDOCU("
                "no_cia,punto,tipo_docu,no_docu,cuenta,centro_costo,"
                "tipo_movi,no_proveedor,monto,afecta_presupuesto"
                ") VALUES(:1,:2,:3,:4,:5,:6,:7,:8,:9,'N')",
                [no_cia, punto, tipo_docu, no_docu,
                 linea.get("cuenta", ""), linea.get("centro_costo", ""),
                 tm, no_prov_num, monto])

        cur.connection.commit()

    return no_docu


def reversar_documento(no_cia, punto, tipo_docu, no_docu, usuario="API", motivo=""):
    """
    Marca documento como reversado (status=R, saldo=0) y genera automaticamente
    el ajuste contrario (AC/AD, TCXP_TDOCU.tipo_transaccion='A') que lo
    contrarresta contablemente -- mismo patron ya validado en
    aplicar_saldos_menores (INSERT TCXP_DOCUMENTO + TCXP_DCDOCU + TCXP_REFEDOCU
    aplicando el ajuste contra el documento original). Sin esto el reverso
    solo tocaba el documento original y no dejaba rastro para 606/mayor.
    Solo sobre docs con status=A. REVERSIBLE en pruebas ZZTEST.
    """
    rows = client.fetch_dicts(
        "SELECT status, no_proveedor, tipo_movi, NVL(saldo,0) AS saldo "
        "FROM CXP.TCXP_DOCUMENTO WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
        [no_cia, punto, tipo_docu, no_docu])
    if not rows:
        raise ValueError("Documento no encontrado")
    doc = rows[0]
    if doc["status"] == "C":
        raise ValueError("Documento ya cerrado, no se puede reversar")
    if doc["status"] == "R":
        raise ValueError("Documento ya está reversado")

    # DETALLE es VARCHAR2(100)
    detalle = ("REVERSADO - " + motivo.strip())[:100] if motivo and motivo.strip() else "REVERSADO"

    # Solo se genera el ajuste por el saldo realmente pendiente. Si el
    # documento ya esta en 0 (pagado/aplicado por otra via aunque status
    # siga en 'A' porque no se ha generado al mayor) no hay nada que
    # contrarrestar y el reverso solo cambia el status.
    monto = round(abs(float(doc["saldo"] or 0)), 2)
    tipo_movi_orig = (doc.get("tipo_movi") or "C").upper()
    # El ajuste debe ser del signo contrario al documento que se reversa:
    # factura (C, aumenta lo que debemos) -> AD ajuste debito (D) la cancela.
    # nota de debito/ajuste (D) -> AC ajuste credito (C) la cancela.
    tipo_movi_ajuste = "D" if tipo_movi_orig == "C" else "C"
    tipo_movi_proveedor = "C" if tipo_movi_ajuste == "D" else "D"

    ajuste = None
    if monto > 0:
        # El ajuste de reverso se emite como Nota de Debito (ND) o Nota de
        # Credito (NC) — es lo que el usuario espera ver contra el proveedor
        # y lo que el reporte 606/mayor consume. Filtrar por
        # tipo_transaccion='A' devolvia AD/AC (AJUSTE DEBITO/CREDITO) que son
        # tipos internos de ajuste y no lo que el usuario reversa.
        tipo_docu_ajuste = "ND" if tipo_movi_ajuste == "D" else "NC"
        tdoc_rows = client.fetch_dicts(
            "SELECT tipo_docu, tipo_transaccion, NVL(cuenta,'') AS cuenta, "
            "NVL(centro_costo,'0000000000') AS centro_costo "
            "FROM CXP.TCXP_TDOCU WHERE tipo_docu=:1 AND tipo_movi=:2",
            [tipo_docu_ajuste, tipo_movi_ajuste])
        if not tdoc_rows:
            raise ValueError(
                f"Falta tipo de documento {tipo_docu_ajuste} configurado en TCXP_TDOCU")
        td = tdoc_rows[0]
        motivo_ajuste = (
            f"{'ND' if tipo_movi_ajuste == 'D' else 'NC'} POR REVERSO "
            f"{tipo_docu}-{no_docu}" + (f" - {motivo.strip()}" if motivo and motivo.strip() else "")
        )[:100]
        with client.cursor() as cur:
            no_docu_ajuste = _next_no_docu(cur, no_cia, punto, td["tipo_docu"])
            _insert_cxp_ajuste_header(
                cur, no_cia, punto, td["tipo_docu"], no_docu_ajuste,
                doc["no_proveedor"], date.today().isoformat(), monto,
                motivo_ajuste, tipo_movi_ajuste, usuario,
                tipo_transaccion=td["tipo_transaccion"])
            _insert_cxp_ajuste_lineas(
                cur, no_cia, punto, td["tipo_docu"], no_docu_ajuste,
                doc["no_proveedor"], monto,
                cuenta_ajuste=td["cuenta"], centro_costo=td["centro_costo"],
                cuenta_proveedor=_CUENTA_CXP_PROVEEDOR_DEFAULT,
                signo_ajuste=tipo_movi_ajuste, signo_proveedor=tipo_movi_proveedor)
            cur.execute(
                "INSERT INTO CXP.TCXP_REFEDOCU "
                "(no_cia, punto, tipo_docu, no_docu, no_proveedor, tipo_refe, no_refe, monto) "
                "VALUES (:1,:2,:3,:4,:5,:6,:7,:8)",
                [no_cia, punto, td["tipo_docu"], no_docu_ajuste, str(doc["no_proveedor"]),
                 tipo_docu, no_docu, monto])
            cur.execute(
                "UPDATE CXP.TCXP_DOCUMENTO SET status='R', saldo=0, detalle=:1, "
                "tipo_docu_r=:2, no_docu_r=:3 "
                "WHERE no_cia=:4 AND punto=:5 AND tipo_docu=:6 AND no_docu=:7",
                [detalle, td["tipo_docu"], no_docu_ajuste, no_cia, punto, tipo_docu, no_docu])
            cur.connection.commit()
        ajuste = {"tipo_docu": td["tipo_docu"], "no_docu": no_docu_ajuste, "monto": monto}
    else:
        with client.cursor() as cur:
            cur.execute(
                "UPDATE CXP.TCXP_DOCUMENTO SET status='R', saldo=0, detalle=:1 "
                "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
                [detalle, no_cia, punto, tipo_docu, no_docu])
            cur.connection.commit()

    result = {"ok": True, "no_docu": no_docu, "status": "R"}
    if ajuste:
        result["nota_debito" if tipo_movi_ajuste == "D" else "nota_credito"] = ajuste
    return result


def liberar_debito(no_cia, punto, no_docu_cr, tipo_docu_cr, debitos):
    """
    Aplica credito (tipo_movi=C) contra lista de debitos (tipo_movi=D).
    - Para docs tipo_movi=C: saldo positivo se reduce (saldo - monto).
    - Para docs tipo_movi=D: saldo negativo se mueve hacia 0 (saldo + monto).
    - El doc credito queda con saldo=0 al final.
    REVERSIBLE: restaurar saldos. Solo probar con docs ZZTEST.
    """
    with client.cursor() as cur:
        for deb in debitos:
            nd    = deb.get("no_docu") or deb.get("no_doc", "")
            td    = deb.get("tipo_docu") or deb.get("tipo_doc", tipo_docu_cr)
            monto = float(deb.get("monto") or deb.get("valor_aplica") or 0)
            # Determine tipo_movi of the debit doc to apply correctly
            deb_rows = client.fetch_dicts(
                "SELECT tipo_movi FROM CXP.TCXP_DOCUMENTO "
                "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
                [no_cia, punto, td, nd])
            tm = deb_rows[0]["tipo_movi"] if deb_rows else "C"
            if tm == "D":
                # saldo is negative; applying payment moves it toward 0
                cur.execute(
                    "UPDATE CXP.TCXP_DOCUMENTO SET saldo=NVL(saldo,0)+:1 "
                    "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
                    [monto, no_cia, punto, td, nd])
            else:
                cur.execute(
                    "UPDATE CXP.TCXP_DOCUMENTO SET saldo=NVL(saldo,0)-:1 "
                    "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
                    [monto, no_cia, punto, td, nd])
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO SET saldo=0 "
            "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
            [no_cia, punto, tipo_docu_cr, no_docu_cr])
        cur.connection.commit()
    return {"ok": True, "aplicaciones": len(debitos)}


def list_documentos_dgii(no_cia, punto, no_proveedor='', tipo_docu='', no_docu=''):
    """
    Fcxp212 — documentos de un proveedor con sus datos DGII (NCF, RNC,
    tipo de gasto, retencion, forma de pago) para la pantalla Corregir NCF.
    """
    conditions = ['d.no_cia=:1', 'd.punto=:2']
    params = [no_cia, punto]
    if no_proveedor:
        params.append(str(no_proveedor))
        conditions.append(f'd.no_proveedor=:{len(params)}')
    if tipo_docu:
        params.append(tipo_docu)
        conditions.append(f'd.tipo_docu=:{len(params)}')
    if no_docu:
        params.append(str(no_docu).strip())
        conditions.append(f'd.no_docu=:{len(params)}')
    where = ' AND '.join(conditions)
    return client.fetch_dicts(f"""
        SELECT * FROM (
            SELECT d.tipo_docu, d.no_docu, d.no_proveedor,
                   NVL(p.nombre,'') AS proveedor,
                   TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
                   NVL(d.valor_original,0) AS valor_original,
                   NVL(d.saldo,0) AS saldo, d.status, d.tipo_movi,
                   d.ncf, d.posiciones_fijas_ncf, NVL(d.rnc,'') AS rnc,
                   NVL(d.impuesto,0) AS impuesto,
                   NVL(d.itbis_retenido,0) AS itbis_retenido,
                   NVL(d.isr_retenido,0) AS isr_retenido,
                   d.tipo_gasto, d.tipo_retencion, d.forma_pago, d.detalle
            FROM CXP.TCXP_DOCUMENTO d
            LEFT JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor = d.no_proveedor
            WHERE {where}
            ORDER BY d.fecha DESC, d.no_docu DESC
        ) WHERE ROWNUM <= 200
    """, params)


def corregir_datos_dgii(d):
    """
    Fcxp212 — corrige NCF y datos DGII de un documento CxP ya registrado.
    No toca valores/saldos: solo ncf, posiciones_fijas_ncf, rnc, impuesto,
    itbis_retenido, isr_retenido, tipo_gasto, tipo_retencion, forma_pago.
    REVERSIBLE: re-ejecutar con los valores anteriores.
    """
    no_cia    = d['no_cia']
    punto     = d.get('punto', '01')
    tipo_docu = d['tipo_docu']
    no_docu   = str(d['no_docu']).strip()

    rows = client.fetch_dicts(
        "SELECT status, no_proveedor, "
        "TO_CHAR(fecha,'YYYY-MM') AS periodo_docu, "
        "ncf, posiciones_fijas_ncf, rnc, impuesto, itbis_retenido, "
        "isr_retenido, tipo_gasto, tipo_retencion, forma_pago "
        "FROM CXP.TCXP_DOCUMENTO "
        "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
        [no_cia, punto, tipo_docu, no_docu])
    if not rows:
        raise ValueError('Documento no encontrado')
    _no_prov_actual = rows[0].get('no_proveedor')
    _antes = {
        "ncf": rows[0].get("ncf"),
        "posiciones_fijas_ncf": rows[0].get("posiciones_fijas_ncf"),
        "rnc": rows[0].get("rnc"),
        "impuesto": rows[0].get("impuesto"),
        "itbis_retenido": rows[0].get("itbis_retenido"),
        "isr_retenido": rows[0].get("isr_retenido"),
        "tipo_gasto": rows[0].get("tipo_gasto"),
        "tipo_retencion": rows[0].get("tipo_retencion"),
        "forma_pago": rows[0].get("forma_pago"),
    }

    # Solo se bloquea corregir NCF/datos DGII de documentos de un periodo YA
    # CERRADO (anterior a TCXP_PUNTO.ano_proceso/mes_proceso) -- ese si ya
    # desalinearia el 606 presentado ante la DGII. Comparar con != tambien
    # bloqueaba documentos del mes calendario actual mientras el punto no
    # habia avanzado su mes_proceso (cierre pendiente), lo cual no es un
    # periodo cerrado.
    _punto_rows = client.fetch_dicts(
        "SELECT ano_proceso, mes_proceso FROM CXP.TCXP_PUNTO "
        "WHERE no_cia=:1 AND punto=:2", [no_cia, punto])
    if _punto_rows:
        _periodo_actual = '{:04d}-{:02d}'.format(
            int(_punto_rows[0]['ano_proceso']), int(_punto_rows[0]['mes_proceso']))
        if rows[0].get('periodo_docu') < _periodo_actual:
            raise ValueError(
                'Este documento pertenece a un periodo contable ya cerrado '
                '({1}); el periodo en curso es {0}.'.format(
                    _periodo_actual, rows[0].get('periodo_docu')))

    _ncf_raw = str(d.get('ncf') or '').strip()
    _ncf_num = int(_ncf_raw) if _ncf_raw.isdigit() else None
    _pos_ncf = (str(d.get('posiciones_fijas_ncf') or d.get('tipo_ncf') or '')
                .strip().upper() or None)
    _tipo_gasto = (str(d.get('tipo_gasto') or '').strip() or None)
    _tipo_ret   = d.get('tipo_retencion')
    _tipo_ret   = int(_tipo_ret) if _tipo_ret not in (None, '', 0) else None
    _forma_pago = d.get('forma_pago')
    _forma_pago = int(_forma_pago) if _forma_pago not in (None, '') else None

    _despues = {
        "ncf": _ncf_num,
        "posiciones_fijas_ncf": _pos_ncf,
        "rnc": d.get('rnc', ''),
        "impuesto": float(d.get('impuesto') or 0),
        "itbis_retenido": float(d.get('itbis_retenido') or 0),
        "isr_retenido": float(d.get('isr_retenido') or 0),
        "tipo_gasto": _tipo_gasto,
        "tipo_retencion": _tipo_ret,
        "forma_pago": _forma_pago,
    }
    from apps.historial.diff import diff_campos
    _cambios = diff_campos(_antes, _despues, etiquetas={
        "ncf": "NCF", "posiciones_fijas_ncf": "Tipo NCF", "rnc": "RNC",
        "impuesto": "ITBIS", "itbis_retenido": "ITBIS retenido",
        "isr_retenido": "ISR retenido", "tipo_gasto": "Tipo de gasto",
        "tipo_retencion": "Tipo de retención", "forma_pago": "Forma de pago",
    })

    with client.cursor() as cur:
        _check_ncf_duplicate(cur, no_cia, _no_prov_actual, _ncf_num, _pos_ncf,
                             exclude=(tipo_docu, no_docu))
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO SET "
            "ncf=:1, posiciones_fijas_ncf=:2, rnc=:3, "
            "impuesto=:4, itbis_retenido=:5, isr_retenido=:6, "
            "tipo_gasto=:7, tipo_retencion=:8, forma_pago=:9 "
            "WHERE no_cia=:10 AND punto=:11 AND tipo_docu=:12 AND no_docu=:13",
            [
                _ncf_num, _pos_ncf, d.get('rnc', ''),
                float(d.get('impuesto') or 0),
                float(d.get('itbis_retenido') or 0),
                float(d.get('isr_retenido') or 0),
                _tipo_gasto, _tipo_ret, _forma_pago,
                no_cia, punto, tipo_docu, no_docu,
            ])
        historial_repo.log_evento(
            cur, usuario=d.get("usuario", "API"), no_cia=no_cia, punto=punto,
            modulo="CXP", tipo_documento=tipo_docu, no_documento=no_docu,
            accion="EDITAR", cambios=_cambios,
        )
        cur.connection.commit()
    return {'ok': True, 'tipo_docu': tipo_docu, 'no_docu': no_docu}


def aplicar_movimientos_pendientes(no_cia, punto, no_proveedor,
                                   tipo_docu_db='', no_docu_db=''):
    """
    Fcxp206 — datos para Aplicación de Movimientos:
    - a_favor: debitos (tipo_movi=D) con saldo a favor (saldo < 0).
    - pendientes: creditos (tipo_movi=C) con saldo, no bloqueados y sin
      aplicacion previa del debito elegido (NOT EXISTS TCXP_REFEDOCU).
    """
    a_favor = client.fetch_dicts(
        "SELECT d.tipo_docu, d.no_docu, TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha, "
        "       NVL(d.valor_original,0) AS valor_original, "
        "       NVL(d.saldo,0) AS saldo, d.detalle "
        "FROM CXP.TCXP_DOCUMENTO d "
        "WHERE d.no_cia=:1 AND d.punto=:2 AND d.no_proveedor=:3 "
        "  AND d.tipo_movi='D' AND NVL(d.saldo,0) < 0 "
        "ORDER BY d.fecha, d.no_docu",
        [no_cia, punto, str(no_proveedor)])

    conditions = [
        "a.no_cia=:1", "a.punto=:2", "a.no_proveedor=:3",
        "a.tipo_movi='C'", "NVL(a.pago_bloqueado,'N')='N'",
        "NVL(a.saldo,0) != 0",
    ]
    params = [no_cia, punto, str(no_proveedor)]
    not_exists = ''
    if tipo_docu_db and no_docu_db:
        params.append(tipo_docu_db)
        p_td = len(params)
        params.append(str(no_docu_db).strip())
        p_nd = len(params)
        not_exists = (
            f" AND NOT EXISTS (SELECT 'x' FROM CXP.TCXP_REFEDOCU b "
            f"WHERE b.no_cia=a.no_cia AND b.punto=a.punto "
            f"AND b.tipo_docu=:{p_td} AND b.no_docu=:{p_nd} "
            f"AND b.tipo_refe=a.tipo_docu AND b.no_refe=a.no_docu)")
    pendientes = client.fetch_dicts(
        "SELECT a.tipo_docu, a.no_docu, TO_CHAR(a.fecha,'YYYY-MM-DD') AS fecha, "
        "       NVL(a.valor_original,0) AS valor_original, "
        "       NVL(a.saldo,0) AS saldo, a.detalle "
        "FROM CXP.TCXP_DOCUMENTO a "
        "WHERE " + ' AND '.join(conditions) + not_exists +
        " ORDER BY a.fecha, a.no_docu",
        params)
    return {'a_favor': a_favor, 'pendientes': pendientes}


def aplicar_movimiento(no_cia, punto, tipo_docu, no_docu, aplicaciones):
    """
    Fcxp206 — aplica un debito con saldo a favor contra facturas (creditos)
    pendientes del mismo proveedor. Por cada aplicacion:
      UPDATE credito saldo=saldo-monto (status=C si queda en 0)
      INSERT TCXP_REFEDOCU (trazabilidad)
      UPDATE debito  saldo=saldo+monto (status=C si queda en 0)
    REVERSIBLE: restaurar saldos y borrar las filas de TCXP_REFEDOCU.
    """
    no_docu = str(no_docu).strip()
    rows = client.fetch_dicts(
        "SELECT NVL(saldo,0) AS saldo, tipo_movi, no_proveedor, status "
        "FROM CXP.TCXP_DOCUMENTO "
        "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
        [no_cia, punto, tipo_docu, no_docu])
    if not rows:
        raise ValueError('Documento a aplicar no encontrado')
    doc = rows[0]
    if (doc['tipo_movi'] or '').upper() != 'D':
        raise ValueError('El documento a aplicar debe ser un débito (saldo a favor)')
    disponible = round(-float(doc['saldo']), 2)
    if disponible <= 0:
        raise ValueError('El documento no tiene saldo a favor disponible')
    no_proveedor = str(doc['no_proveedor']).strip()

    total = round(sum(float(a.get('monto') or 0) for a in aplicaciones), 2)
    if total <= 0:
        raise ValueError('No hay montos a aplicar')
    if total > disponible + 0.005:
        raise ValueError(
            f'El total a aplicar ({total:,.2f}) excede el saldo a favor '
            f'disponible ({disponible:,.2f})')

    aplicados = []
    with client.cursor() as cur:
        saldo_db = float(doc['saldo'])
        for ap in aplicaciones:
            monto    = round(float(ap.get('monto') or 0), 2)
            if monto <= 0:
                continue
            tipo_ref = (ap.get('tipo_docu') or ap.get('tipo_refe') or '').strip()
            no_ref   = str(ap.get('no_docu') or ap.get('no_refe') or '').strip()
            ref_rows = client.fetch_dicts(
                "SELECT NVL(saldo,0) AS saldo, tipo_movi, "
                "       NVL(pago_bloqueado,'N') AS pago_bloqueado "
                "FROM CXP.TCXP_DOCUMENTO "
                "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
                [no_cia, punto, tipo_ref, no_ref])
            if not ref_rows:
                raise ValueError(f'Factura {tipo_ref}-{no_ref} no encontrada')
            ref = ref_rows[0]
            if (ref['tipo_movi'] or '').upper() != 'C':
                raise ValueError(
                    f'{tipo_ref}-{no_ref} no es un crédito pendiente de pago')
            if ref['pago_bloqueado'] == 'S':
                raise ValueError(f'{tipo_ref}-{no_ref} tiene el pago bloqueado')
            saldo_cr = float(ref['saldo'])
            if monto > saldo_cr + 0.005:
                raise ValueError(
                    f'El monto a aplicar a {tipo_ref}-{no_ref} ({monto:,.2f}) '
                    f'excede su saldo ({saldo_cr:,.2f})')

            nuevo_cr = round(saldo_cr - monto, 2)
            st_cr = 'C' if abs(nuevo_cr) < 0.005 else None
            if st_cr:
                cur.execute(
                    "UPDATE CXP.TCXP_DOCUMENTO SET saldo=0, status='C' "
                    "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
                    [no_cia, punto, tipo_ref, no_ref])
            else:
                cur.execute(
                    "UPDATE CXP.TCXP_DOCUMENTO SET saldo=:1 "
                    "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
                    [nuevo_cr, no_cia, punto, tipo_ref, no_ref])
            cur.execute(
                "INSERT INTO CXP.TCXP_REFEDOCU "
                "(no_cia, punto, tipo_docu, no_docu, no_proveedor, "
                " tipo_refe, no_refe, monto) "
                "VALUES (:1,:2,:3,:4,:5,:6,:7,:8)",
                [no_cia, punto, tipo_docu, no_docu, no_proveedor,
                 tipo_ref, no_ref, monto])
            saldo_db = round(saldo_db + monto, 2)
            aplicados.append({'tipo_docu': tipo_ref, 'no_docu': no_ref,
                              'monto': monto, 'saldo_restante': nuevo_cr})

        if not aplicados:
            raise ValueError('No hay montos a aplicar')
        if abs(saldo_db) < 0.005:
            cur.execute(
                "UPDATE CXP.TCXP_DOCUMENTO SET saldo=0, status='C' "
                "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
                [no_cia, punto, tipo_docu, no_docu])
        else:
            cur.execute(
                "UPDATE CXP.TCXP_DOCUMENTO SET saldo=:1 "
                "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
                [saldo_db, no_cia, punto, tipo_docu, no_docu])
        cur.connection.commit()

    return {'ok': True, 'aplicaciones': aplicados,
            'saldo_favor_restante': saldo_db}


def bloquear_pago(no_cia, punto, tipo_docu, no_docu, bloquear=True):
    """Toggle pago_bloqueado S/N. REVERSIBLE: llamar de nuevo con bloquear=False."""
    rows = client.fetch_dicts(
        "SELECT no_docu FROM CXP.TCXP_DOCUMENTO "
        "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
        [no_cia, punto, tipo_docu, no_docu])
    if not rows:
        raise ValueError("Documento no encontrado")
    flag = "S" if bloquear else "N"
    with client.cursor() as cur:
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO SET pago_bloqueado=:1 "
            "WHERE no_cia=:2 AND punto=:3 AND tipo_docu=:4 AND no_docu=:5",
            [flag, no_cia, punto, tipo_docu, no_docu])
        cur.connection.commit()
    return {"ok": True, "pago_bloqueado": flag}


# ─── SOLICITUDES DE PAGO (Fcxp209 / Fcxp207 — puente CxP → CHC) ──────────────


def documentos_por_pagar(no_cia, punto, no_proveedor):
    """Documentos del proveedor con saldo > 0 (tipo_movi='C') candidatos a
    solicitud de cheque. monto_solicitado = lo ya comprometido en solicitudes
    SO pendientes (TCHC_CHEQUE tipo_transaccion='K', no impresas, activas) —
    misma consulta que usa el legado en Fcxp207/209."""
    return client.fetch_dicts(
        "SELECT d.tipo_docu, d.no_docu, d.no_proveedor, d.fecha, d.fecha_vence, "
        "       d.valor_original, d.saldo, d.pago_bloqueado, d.detalle, "
        "       NVL((SELECT SUM(NVL(r.monto,0)) "
        "              FROM CHC.TCHC_CHEQUE c, CHC.TCHC_REFEDOCU r "
        "             WHERE c.no_cia=d.no_cia AND c.punto=d.punto "
        "               AND c.tipo_transaccion='K' AND c.st_impresion='N' "
        "               AND c.st_nulo='A' "
        "               AND r.no_cia=c.no_cia AND r.punto=c.punto "
        "               AND r.tipo_docu=c.tipo_docu AND r.no_docu=c.no_docu "
        "               AND r.tipo_refe=d.tipo_docu AND r.no_refe=d.no_docu "
        "               AND r.no_proveedor=d.no_proveedor), 0) monto_solicitado "
        "  FROM CXP.TCXP_DOCUMENTO d "
        " WHERE d.no_cia=:1 AND d.punto=:2 AND d.no_proveedor=:3 "
        "   AND d.tipo_movi='C' AND NVL(d.saldo,0) > 0 "
        " ORDER BY d.fecha, d.no_docu",
        [no_cia, punto, no_proveedor],
    )


def generar_solicitud_cheque(no_cia, punto, cuenta_banco, no_proveedor,
                             docs, usuario, fecha_cheque=None, detalle=None):
    """Fcxp209 — Generar Solicitud a Cheque.

    Crea un documento SO (Solicitud de Cheque) en CHC a partir de documentos
    CxP con saldo del proveedor, en una sola transacción:
    - TCHC_CHEQUE tipo_docu='SO' (tipo_movi='C', tipo_transaccion='K') en el
      estado pendiente exacto del legado: status='N', que_es='S',
      autorizado='N', st_impresion='N', st_nulo='A'.
    - Una fila TCHC_REFEDOCU por cada documento CxP aplicado.
    - TCHC_BCUENTA.che_por_entregar += total.
    No toca TCXP_DOCUMENTO.saldo: el saldo se aplica cuando CHC emite el
    cheque. (TCXP_SOLICITUD era la mesa de selección temporal del Forms; la
    selección vive ahora en el cliente y no se persiste.)

    docs: [{tipo_docu, no_docu, monto}]
    """
    if not docs:
        raise ValueError('Debe indicar al menos un documento a pagar')

    prov = client.fetch_one(
        "SELECT p.nombre, NVL(b.activo,'S') activo "
        "  FROM CXP.TCXP_DPROVEEDOR p, CXP.TCXP_BPROVEEDOR b "
        " WHERE b.no_cia=:1 AND b.punto=:2 AND b.no_proveedor=:3 "
        "   AND p.no_proveedor=b.no_proveedor",
        [no_cia, punto, no_proveedor],
    )
    if not prov:
        raise ValueError(f'Proveedor {no_proveedor} no existe en la empresa/punto')
    if prov[1] != 'S':
        raise ValueError(f'Proveedor {no_proveedor} está inactivo')
    beneficiario = (prov[0] or '').strip().upper()

    bcuenta = client.fetch_one(
        "SELECT moneda, NVL(activa,'S') activa FROM CHC.TCHC_BCUENTA "
        " WHERE no_cia=:1 AND punto=:2 AND cuenta_banco=:3",
        [no_cia, punto, cuenta_banco],
    )
    if not bcuenta:
        raise ValueError(f'Cuenta {cuenta_banco} no existe en empresa {no_cia} punto {punto}')
    if bcuenta[1] != 'S':
        raise ValueError(f'Cuenta {cuenta_banco} está inactiva')
    moneda = bcuenta[0] or 'P'

    with client.connection() as conn:
        cur = conn.cursor()
        total = 0.0
        refs = []
        for d in docs:
            td = (d.get('tipo_docu') or '').strip()
            nd = (d.get('no_docu') or '').strip()
            monto = round(float(d.get('monto') or 0), 2)
            if monto <= 0:
                raise ValueError(f'Monto inválido para {td}-{nd}')
            cur.execute(
                "SELECT NVL(saldo,0), NVL(pago_bloqueado,'N') "
                "  FROM CXP.TCXP_DOCUMENTO "
                " WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4 "
                "   AND tipo_movi='C' FOR UPDATE",
                [no_cia, punto, td, nd])
            row = cur.fetchone()
            if not row:
                raise ValueError(f'Documento {td}-{nd} no existe o no es por pagar')
            saldo, bloqueado = float(row[0]), row[1]
            if bloqueado == 'S':
                raise ValueError(f'Documento {td}-{nd} tiene el pago bloqueado')
            cur.execute(
                "SELECT NVL(SUM(NVL(r.monto,0)),0) "
                "  FROM CHC.TCHC_CHEQUE c, CHC.TCHC_REFEDOCU r "
                " WHERE c.no_cia=:1 AND c.punto=:2 AND c.tipo_transaccion='K' "
                "   AND c.st_impresion='N' AND c.st_nulo='A' "
                "   AND r.no_cia=c.no_cia AND r.punto=c.punto "
                "   AND r.tipo_docu=c.tipo_docu AND r.no_docu=c.no_docu "
                "   AND r.tipo_refe=:3 AND r.no_refe=:4 AND r.no_proveedor=:5",
                [no_cia, punto, td, nd, no_proveedor])
            ya_solicitado = float(cur.fetchone()[0])
            disponible = round(saldo - ya_solicitado, 2)
            if monto > disponible:
                raise ValueError(
                    f'{td}-{nd}: el monto {monto:,.2f} excede el disponible '
                    f'{disponible:,.2f} (saldo {saldo:,.2f} menos '
                    f'{ya_solicitado:,.2f} ya en solicitudes pendientes)')
            total = round(total + monto, 2)
            refs.append((td, nd, monto))

        # Correlativo SO desde TCHC_SECUENCIA (misma transacción)
        cur.execute(
            "SELECT NVL(ult_docu,0) FROM CHC.TCHC_SECUENCIA "
            " WHERE no_cia=:1 AND punto=:2 AND tipo_docu='SO' FOR UPDATE",
            [no_cia, punto])
        row = cur.fetchone()
        if row is None:
            cur.execute(
                "INSERT INTO CHC.TCHC_SECUENCIA (no_cia, punto, tipo_docu, ult_docu) "
                "VALUES (:1,:2,'SO',1)", [no_cia, punto])
            no_docu_so = '1'.zfill(7)
        else:
            nxt = int(row[0]) + 1
            cur.execute(
                "UPDATE CHC.TCHC_SECUENCIA SET ult_docu=:1 "
                " WHERE no_cia=:2 AND punto=:3 AND tipo_docu='SO'",
                [nxt, no_cia, punto])
            no_docu_so = str(nxt).zfill(7)

        cur.execute(
            "INSERT INTO CHC.TCHC_CHEQUE ("
            "  no_cia, punto, tipo_docu, no_docu, cuenta_banco, no_proveedor, "
            "  tipo_movi, tipo_transaccion, fecha_solicitud, fecha_cheque, "
            "  beneficiario, status, st_impresion, st_nulo, "
            "  autorizado, conciliado, entregado, retenido, que_es, deposito_tc, "
            "  debito, credito, valor_original, saldo, total_refe, usuario, "
            "  moneda_cuenta, pago, st_generado_cnt, detalle1, estado_encf, "
            "  fecha_sysdate"
            ") VALUES ("
            "  :1, :2, 'SO', :3, :4, :5, "
            "  'C', 'K', SYSDATE, "
            "  CASE WHEN :6 IS NULL THEN SYSDATE ELSE TO_DATE(:7,'YYYY-MM-DD') END, "
            "  :8, 'N', 'N', 'A', "
            "  'N', 'N', 'N', 'N', 'S', 'N', "
            "  :9, :10, :11, :12, :13, :14, "
            "  :15, 'N', 'N', :16, 0, SYSDATE"
            ")",
            [no_cia, punto, no_docu_so, cuenta_banco, no_proveedor,
             fecha_cheque, fecha_cheque, beneficiario,
             total, total, total, total, total, usuario,
             moneda, (detalle or '').strip() or None])

        for td, nd, monto in refs:
            cur.execute(
                "INSERT INTO CHC.TCHC_REFEDOCU "
                "  (no_cia, punto, tipo_docu, no_docu, tipo_refe, no_refe, "
                "   no_proveedor, monto) "
                "VALUES (:1,:2,'SO',:3,:4,:5,:6,:7)",
                [no_cia, punto, no_docu_so, td, nd, no_proveedor, monto])

        cur.execute(
            "UPDATE CHC.TCHC_BCUENTA "
            "   SET che_por_entregar = NVL(che_por_entregar,0) + :1 "
            " WHERE no_cia=:2 AND punto=:3 AND cuenta_banco=:4",
            [total, no_cia, punto, cuenta_banco])

        conn.commit()

    return {
        'ok': True,
        'tipo_docu': 'SO',
        'no_docu': no_docu_so,
        'beneficiario': beneficiario,
        'total': total,
        'documentos': len(refs),
    }


def solicitudes_pago(no_cia, punto, no_proveedor=None, pendientes='S', limit=200):
    """Fcxp207 — Procesar/consultar Solicitudes de Pago.

    Lista las solicitudes de cheque (TCHC_CHEQUE tipo_transaccion='K').
    pendientes='S' → solo no impresas y activas (estado pendiente legado)."""
    sql = (
        "SELECT * FROM ("
        "SELECT c.tipo_docu, c.no_docu, c.no_proveedor, c.beneficiario, "
        "       c.cuenta_banco, c.fecha_solicitud, c.fecha_cheque, "
        "       c.valor_original, c.status, c.st_impresion, c.st_nulo, "
        "       c.autorizado, c.que_es, c.usuario, c.detalle1, c.no_cheque "
        "  FROM CHC.TCHC_CHEQUE c "
        " WHERE c.no_cia=:no_cia AND c.punto=:punto "
        "   AND c.tipo_transaccion='K' "
    )
    params = {'no_cia': no_cia, 'punto': punto}
    if no_proveedor:
        sql += " AND c.no_proveedor=:no_proveedor "
        params['no_proveedor'] = no_proveedor
    if pendientes == 'S':
        sql += " AND c.st_impresion='N' AND c.st_nulo='A' "
    sql += (" ORDER BY c.fecha_solicitud DESC, c.no_docu DESC"
            ") WHERE ROWNUM <= :lim")
    params['lim'] = int(limit)
    return client.fetch_dicts(sql, params)


def solicitud_referencias(no_cia, punto, no_docu):
    """Documentos CxP referenciados por una solicitud SO (TCHC_REFEDOCU)."""
    return client.fetch_dicts(
        "SELECT r.tipo_refe, r.no_refe, r.no_proveedor, NVL(r.monto,0) monto, "
        "       d.fecha, d.detalle, d.valor_original, d.saldo "
        "  FROM CHC.TCHC_REFEDOCU r, CXP.TCXP_DOCUMENTO d "
        " WHERE r.no_cia=:1 AND r.punto=:2 AND r.tipo_docu='SO' AND r.no_docu=:3 "
        "   AND d.no_cia(+)=r.no_cia AND d.punto(+)=r.punto "
        "   AND d.tipo_docu(+)=r.tipo_refe AND d.no_docu(+)=r.no_refe "
        " ORDER BY r.tipo_refe, r.no_refe",
        [no_cia, punto, no_docu],
    )


# ─── CIERRE / ASIENTO (IRREVERSIBLES - construidos, NO ejecutar sin supervision) ─


def get_asiento_contable_cxp(no_cia, punto, mes, ano):
    """Consolida lineas TCXP_DCDOCU para el mes/año dado (read-only, seguro)."""
    return client.fetch_dicts(
        "SELECT l.cuenta, l.centro_costo, l.tipo_movi, "
        "SUM(CASE WHEN l.tipo_movi='D' THEN NVL(l.monto,0) ELSE 0 END) total_debito, "
        "SUM(CASE WHEN l.tipo_movi='C' THEN NVL(l.monto,0) ELSE 0 END) total_credito "
        "FROM CXP.TCXP_DCDOCU l "
        "JOIN CXP.TCXP_DOCUMENTO d "
        "  ON d.no_cia=l.no_cia AND d.punto=l.punto "
        "  AND d.tipo_docu=l.tipo_docu AND d.no_docu=l.no_docu "
        "WHERE l.no_cia=:1 AND d.punto=:2 "
        "AND EXTRACT(MONTH FROM d.fecha)=:3 AND EXTRACT(YEAR FROM d.fecha)=:4 "
        "AND d.status='A' "
        "GROUP BY l.cuenta, l.centro_costo, l.tipo_movi ORDER BY l.cuenta",
        [no_cia, punto, mes, ano])


def generar_asiento_mayor_cxp(no_cia, punto, mes_proceso, ano_proceso):
    """
    Marca docs del periodo como generados al mayor (st_generado_cnt=S).
    IRREVERSIBLE sobre datos reales. NO ejecutar sin supervision.
    """
    with client.cursor() as cur:
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO SET st_generado_cnt='S' "
            "WHERE no_cia=:1 AND punto=:2 "
            "AND EXTRACT(MONTH FROM fecha)=:3 AND EXTRACT(YEAR FROM fecha)=:4 "
            "AND status='A' AND NVL(st_generado_cnt,'N')='N'",
            [no_cia, punto, mes_proceso, ano_proceso])
        cur.connection.commit()
    return {"ok": True}


def cierre_cxp(no_cia, punto):
    """
    Avanza mes_proceso en TCXP_PUNTO al mes siguiente.
    IRREVERSIBLE sobre datos reales. NO ejecutar sin supervision.
    """
    row = client.fetch_one(
        "SELECT NVL(mes_proceso,1), NVL(ano_proceso,2025) "
        "FROM CXP.TCXP_PUNTO WHERE no_cia=:1 AND punto=:2",
        [no_cia, punto])
    if not row:
        raise ValueError("Punto no encontrado")
    mes, ano = row
    if mes == 12:
        nuevo_mes, nuevo_ano = 1, ano + 1
    else:
        nuevo_mes, nuevo_ano = mes + 1, ano
    with client.cursor() as cur:
        cur.execute(
            "UPDATE CXP.TCXP_PUNTO SET mes_proceso=:1, ano_proceso=:2 "
            "WHERE no_cia=:3 AND punto=:4",
            [nuevo_mes, nuevo_ano, no_cia, punto])
        cur.connection.commit()
    # Al abrir el nuevo mes, materializar los documentos que estaban en cola
    # esperando exactamente ese periodo (compras/facturas de INV/FAT o manuales
    # capturadas por adelantado). No rompe el cierre si alguno falla.
    materializados = _materializar_cola(no_cia, punto, nuevo_ano, nuevo_mes)
    return {"mes": nuevo_mes, "ano": nuevo_ano, "materializados": materializados}


def _materializar_cola(no_cia: str, punto: str, ano: int, mes: int) -> list[dict]:
    """Inserta en TCXP_DOCUMENTO los documentos en cola cuyo periodo objetivo
    es (ano, mes) y estan PENDIENTE. Reusa entrada_documento (con el candado
    de periodo desactivado) para replicar exactamente la logica de creacion
    (NCF, secuencia, partida doble). Cada doc se procesa aislado: un error
    marca esa fila como ERROR y sigue con las demas."""
    # payload es CLOB: leerlo como str DENTRO del cursor abierto (fetch_dicts
    # lo devolveria como LOB atado a un cursor ya cerrado).
    rows = []
    with client.cursor() as cur:
        cur.execute(
            "SELECT id, payload FROM CXP.TCXP_DOCUMENTO_COLA "
            "WHERE no_cia=:1 AND punto=:2 AND ano_objetivo=:3 AND mes_objetivo=:4 "
            "AND estado='PENDIENTE' ORDER BY id",
            [no_cia, punto, ano, mes])
        for cid, payload_lob in cur.fetchall():
            payload_str = payload_lob.read() if hasattr(payload_lob, 'read') else payload_lob
            rows.append({'id': cid, 'payload': payload_str})
    res = []
    for r in rows:
        cid = r['id']
        try:
            payload = json.loads(r['payload'])
            link = payload.pop('_cola_link', None)
            payload.pop('_cola_origen', None)
            no_docu = entrada_documento(payload, _skip_periodo_gate=True)
            if link and no_docu:
                # Enlace diferido EC(INV) -> FT(CxP): completa TINV_RME.no_refe
                # que quedo vacio cuando el espejo se encolo. tipo_refe apunta al
                # tipo del documento espejo creado (FT), no al EC de origen.
                try:
                    client.execute(
                        "UPDATE INV.TINV_RME SET tipo_refe=:1, no_refe=:2 "
                        "WHERE no_cia=:3 AND punto=:4 AND tipo_docu=:5 AND no_docu=:6",
                        [payload.get('tipo_docu', 'FT'), str(no_docu),
                         link['no_cia'], link['punto'],
                         link.get('tipo_docu', 'EC'), link['no_docu']])
                except Exception:
                    pass
            client.execute(
                "UPDATE CXP.TCXP_DOCUMENTO_COLA SET estado='MATERIALIZADO', "
                "no_docu_generado=:1, fecha_materializado=SYSDATE, mensaje_error=NULL "
                "WHERE id=:2", [str(no_docu), cid])
            res.append({'id': cid, 'no_docu': no_docu, 'ok': True})
        except Exception as exc:
            try:
                client.execute(
                    "UPDATE CXP.TCXP_DOCUMENTO_COLA SET estado='ERROR', "
                    "mensaje_error=:1 WHERE id=:2", [str(exc)[:500], cid])
            except Exception:
                pass
            res.append({'id': cid, 'error': str(exc)})
    return res


def listar_cola(no_cia: str, punto: str = '', estado: str = '') -> dict:
    """Lista los documentos en cola para la pantalla de revision."""
    conds = ["no_cia=:1"]; params: list = [no_cia]
    if punto:
        params.append(punto); conds.append('punto=:' + str(len(params)))
    if estado:
        params.append(estado.upper()); conds.append('estado=:' + str(len(params)))
    where = ' AND '.join(conds)
    rows = client.fetch_dicts(
        "SELECT id, no_cia, punto, ano_objetivo, mes_objetivo, origen, "
        "tipo_docu, no_proveedor, ncf, TO_CHAR(fecha_doc,'YYYY-MM-DD') fecha_doc, "
        "valor, estado, usuario, TO_CHAR(fecha_encolado,'YYYY-MM-DD HH24:MI') fecha_encolado, "
        "TO_CHAR(fecha_materializado,'YYYY-MM-DD HH24:MI') fecha_materializado, "
        "no_docu_generado, mensaje_error "
        "FROM CXP.TCXP_DOCUMENTO_COLA WHERE " + where +
        " ORDER BY estado, ano_objetivo, mes_objetivo, id", params)
    pend = sum(1 for r in rows if r['estado'] == 'PENDIENTE')
    return {'items': rows, 'count': len(rows), 'pendientes': pend}


def materializar_cola_manual(cid: int) -> dict:
    """Materializa una fila de cola bajo demanda -- solo si su periodo objetivo
    ya es <= el periodo de proceso abierto (es decir, ese mes ya se abrio)."""
    r = client.fetch_one(
        "SELECT no_cia, punto, ano_objetivo, mes_objetivo, estado "
        "FROM CXP.TCXP_DOCUMENTO_COLA WHERE id=:1", [cid])
    if not r:
        raise ValueError("Fila de cola no encontrada")
    no_cia, punto, ano_obj, mes_obj, estado = r
    if estado != 'PENDIENTE':
        raise ValueError(f"La fila ya esta {estado}, no PENDIENTE")
    prow = client.fetch_one(
        "SELECT NVL(ano_proceso,0), NVL(mes_proceso,0) FROM CXP.TCXP_PUNTO "
        "WHERE no_cia=:1 AND punto=:2", [no_cia, punto])
    per_obj = ano_obj * 100 + mes_obj
    per_proc = int(prow[0]) * 100 + int(prow[1]) if prow else 0
    if per_obj > per_proc:
        raise ValueError(
            'El periodo {0:04d}-{1:02d} aun no esta abierto en CxP (proceso en '
            '{2:04d}-{3:02d}). Cierre/avance el periodo primero.'.format(
                ano_obj, mes_obj, int(prow[0]), int(prow[1])))
    out = _materializar_cola(no_cia, punto, ano_obj, mes_obj)
    return {'ok': True, 'resultado': [x for x in out]}


def get_cola_documento(cid: int) -> dict | None:
    """Payload completo de una fila de cola PENDIENTE, en la misma forma que
    espera el formulario de Entrada de Documentos (editTipo/editNoDocu) --
    permite reabrir y corregir un documento antes de que se materialice,
    igual que el boton Editar de Consulta de Documentos hace con uno ya
    creado. No aplica a filas ya MATERIALIZADO/ERROR (esas se editan sobre
    el documento real via entrada_documento)."""
    with client.cursor() as cur:
        cur.execute(
            "SELECT estado, payload FROM CXP.TCXP_DOCUMENTO_COLA WHERE id=:1", [cid])
        row = cur.fetchone()
        if not row:
            return None
        estado, payload_lob = row
        payload_str = payload_lob.read() if hasattr(payload_lob, 'read') else payload_lob
    if estado != 'PENDIENTE':
        raise ValueError(f"La fila ya esta {estado}, no se puede editar")
    payload = json.loads(payload_str)
    payload.pop('_cola_link', None)
    payload.pop('_cola_origen', None)
    prov = client.fetch_dicts(
        "SELECT nombre, rnc AS rnc_ficha, direccion FROM CXP.TCXP_DPROVEEDOR "
        "WHERE no_proveedor=:1", [str(payload.get('no_proveedor') or '')])
    doc = dict(payload)
    doc['cola_id'] = cid
    doc['nombre_proveedor'] = prov[0]['nombre'] if prov else ''
    doc['direccion_proveedor'] = prov[0]['direccion'] if prov else ''
    doc['posiciones_fijas_ncf'] = payload.get('tipo_ncf') or payload.get('posiciones_fijas_ncf')
    doc['valor_original'] = payload.get('valor_original') or payload.get('valor') or 0
    doc['debitos_aplicados'] = []
    return doc


def editar_cola_documento(cid: int, d: dict) -> dict:
    """Actualiza el payload guardado de una fila PENDIENTE de la cola --
    mismas validaciones minimas que entrada_documento hace antes de encolar
    (fecha valida, lineas balanceadas), pero sin tocar TCXP_DOCUMENTO: la
    fila sigue PENDIENTE hasta que su periodo se materialice."""
    with client.cursor() as cur:
        cur.execute(
            "SELECT estado, no_cia, punto FROM CXP.TCXP_DOCUMENTO_COLA WHERE id=:1", [cid])
        row = cur.fetchone()
        if not row:
            raise ValueError("Fila de cola no encontrada")
        estado, no_cia, punto = row
        if estado != 'PENDIENTE':
            raise ValueError(f"La fila ya esta {estado}, no se puede editar")

        fecha = _norm_fecha(d.get('fecha', ''), campo='fecha', requerido=True)
        lineas_in = d.get('lineas', [])
        if not lineas_in:
            raise ValueError(
                "El documento debe incluir la distribucion contable "
                "(lineas de Debito/Credito) -- no se puede grabar sin ella.")
        tot_deb = sum(float(l.get('monto') or 0) for l in lineas_in
                      if (l.get('tipo_movi') or 'D').upper() == 'D')
        tot_cre = sum(float(l.get('monto') or 0) for l in lineas_in
                      if (l.get('tipo_movi') or 'D').upper() == 'C')
        if round(abs(tot_deb - tot_cre), 2) > 0.01:
            raise ValueError(
                "La distribucion contable no cuadra: Debito {:.2f} vs "
                "Credito {:.2f}.".format(tot_deb, tot_cre))

        ano_obj, mes_obj = int(fecha[:4]), int(fecha[5:7])
        payload = {**d, 'no_cia': no_cia, 'punto': punto, 'fecha': fecha}
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO_COLA SET "
            "ano_objetivo=:1, mes_objetivo=:2, tipo_docu=:3, no_proveedor=:4, "
            "ncf=:5, fecha_doc=TO_DATE(:6,'YYYY-MM-DD'), valor=:7, payload=:8 "
            "WHERE id=:9",
            [ano_obj, mes_obj, d.get('tipo_docu'), str(d.get('no_proveedor') or ''),
             str(d.get('ncf') or '')[:30], fecha,
             float(d.get('valor_original') or d.get('valor') or 0),
             json.dumps(payload, default=str), cid])
        cur.connection.commit()
    return {'ok': True, 'id': cid}


# ─── SALDOS MENORES POR AJUSTAR (Fcxp204) ──────────────────────────────────

def get_max_saldo_menor_aj(no_cia: str, punto: str) -> float:
    row = client.fetch_one(
        "SELECT NVL(max_saldo_menor_aj,0) FROM CXP.TCXP_PUNTO "
        "WHERE no_cia=:1 AND punto=:2",
        [no_cia, punto])
    return float(row[0]) if row else 0.0


def get_saldos_menores_preview(no_cia: str, punto: str, max_saldo: float):
    """Preview de docs CxP con saldo en (-max_saldo, +max_saldo)\\{0}.
    Positivos → AC (Crédito), Negativos → AD (Débito)."""
    if max_saldo <= 0:
        return {'positivos': [], 'negativos': [], 'max_saldo': max_saldo}
    sql = """
        SELECT d.no_proveedor, NVL(p.nombre,'') AS nombre_proveedor,
               d.tipo_docu AS tipo_doc, d.no_docu AS no_doc,
               TO_CHAR(d.fecha,'YYYY-MM-DD') AS fecha,
               NVL(d.valor_original,0) AS valor,
               NVL(d.saldo,0) AS saldo, d.ncf
          FROM CXP.TCXP_DOCUMENTO d
          LEFT JOIN CXP.TCXP_DPROVEEDOR p ON p.no_proveedor=d.no_proveedor
         WHERE d.no_cia=:1 AND d.punto=:2
           AND d.status='A'
           AND ((d.saldo > 0 AND d.saldo <= :3)
                OR (d.saldo < 0 AND d.saldo >= :4))
         ORDER BY d.no_proveedor, d.fecha, d.no_docu
    """
    rows = client.fetch_dicts(sql, [no_cia, punto, max_saldo, -max_saldo])
    pos: dict = {}
    neg: dict = {}
    for r in rows:
        bucket = pos if r['saldo'] > 0 else neg
        prov = str(r['no_proveedor'] or '').strip()
        if prov not in bucket:
            bucket[prov] = {
                'no_proveedor': prov,
                'nombre_proveedor': r['nombre_proveedor'],
                'docs': [], 'total_saldo': 0.0,
            }
        bucket[prov]['docs'].append({
            'tipo_doc': r['tipo_doc'].strip(),
            'no_doc': r['no_doc'].strip(),
            'fecha': r['fecha'],
            'valor': float(r['valor']),
            'saldo': float(r['saldo']),
            'ncf': r['ncf'],
        })
        bucket[prov]['total_saldo'] += float(r['saldo'])
    return {
        'max_saldo': max_saldo,
        'positivos': sorted(pos.values(), key=lambda x: -x['total_saldo']),
        'negativos': sorted(neg.values(), key=lambda x: x['total_saldo']),
    }


# Cuenta CxP estándar del proveedor (Cuentas por Pagar - Proveedores Locales)
_CUENTA_CXP_PROVEEDOR_DEFAULT = '2101-01'


def aplicar_saldos_menores(no_cia: str, punto: str, max_saldo: float,
                           fecha: str, motivo: str = '',
                           usuario: str = '') -> dict:
    if max_saldo <= 0:
        raise ValueError("max_saldo debe ser mayor que 0")
    preview = get_saldos_menores_preview(no_cia, punto, max_saldo)
    motivo = motivo or 'DOC. GENERADO POR SALDOS MENORES POR AJUSTAR'

    tdoc_rows = client.fetch_dicts(
        "SELECT tipo_docu, tipo_movi, NVL(cuenta,'') AS cuenta, "
        "NVL(centro_costo,'0000000000') AS centro_costo "
        "FROM CXP.TCXP_TDOCU WHERE tipo_transaccion='A'",
        [])
    tdocs = {(r['tipo_movi'] or '').upper(): r for r in tdoc_rows}
    if 'C' not in tdocs or 'D' not in tdocs:
        raise ValueError("Faltan tipos AC/AD configurados en TCXP_TDOCU.")

    docs_creados = []
    with client.cursor() as cur:
        td_ac = tdocs['C']
        for grp in preview['positivos']:
            no_prov = grp['no_proveedor']
            total = round(grp['total_saldo'], 2)
            if total <= 0:
                continue
            no_doc = _next_no_docu(cur, no_cia, punto, td_ac['tipo_docu'])
            _insert_cxp_ajuste_header(cur, no_cia, punto, td_ac['tipo_docu'],
                                       no_doc, no_prov, fecha, total, motivo, 'C',
                                       usuario)
            _insert_cxp_ajuste_lineas(cur, no_cia, punto, td_ac['tipo_docu'],
                                       no_doc, no_prov, total,
                                       cuenta_ajuste=td_ac['cuenta'],
                                       centro_costo=td_ac['centro_costo'],
                                       cuenta_proveedor=_CUENTA_CXP_PROVEEDOR_DEFAULT,
                                       signo_ajuste='C',
                                       signo_proveedor='D')
            _aplicar_refedocu_cxp(cur, no_cia, punto, td_ac['tipo_docu'],
                                  no_doc, no_prov, grp['docs'])
            docs_creados.append({
                'tipo_doc': td_ac['tipo_docu'], 'no_doc': no_doc,
                'no_proveedor': no_prov, 'total': total,
                'cancelados': len(grp['docs']),
            })

        td_ad = tdocs['D']
        for grp in preview['negativos']:
            no_prov = grp['no_proveedor']
            total = round(abs(grp['total_saldo']), 2)
            if total <= 0:
                continue
            no_doc = _next_no_docu(cur, no_cia, punto, td_ad['tipo_docu'])
            _insert_cxp_ajuste_header(cur, no_cia, punto, td_ad['tipo_docu'],
                                       no_doc, no_prov, fecha, total, motivo, 'D',
                                       usuario)
            _insert_cxp_ajuste_lineas(cur, no_cia, punto, td_ad['tipo_docu'],
                                       no_doc, no_prov, total,
                                       cuenta_ajuste=td_ad['cuenta'],
                                       centro_costo=td_ad['centro_costo'],
                                       cuenta_proveedor=_CUENTA_CXP_PROVEEDOR_DEFAULT,
                                       signo_ajuste='D',
                                       signo_proveedor='C')
            _aplicar_refedocu_cxp(cur, no_cia, punto, td_ad['tipo_docu'],
                                  no_doc, no_prov, grp['docs'])
            docs_creados.append({
                'tipo_doc': td_ad['tipo_docu'], 'no_doc': no_doc,
                'no_proveedor': no_prov, 'total': total,
                'cancelados': len(grp['docs']),
            })

        cur.connection.commit()
    return {
        'ok': True,
        'docs_creados': docs_creados,
        'proveedores_positivos': len(preview['positivos']),
        'proveedores_negativos': len(preview['negativos']),
    }


def _insert_cxp_ajuste_header(cur, no_cia, punto, tipo_docu, no_docu,
                               no_proveedor, fecha, valor, motivo, tipo_movi,
                               usuario, tipo_transaccion='A'):
    # CK_TCXPDOCU_DEBITO_CREDITO exige debito=credito=valor_original siempre
    # (igual que entrada_documento) -- la direccion D/C la marca tipo_movi,
    # no poner 0 en el lado contrario.
    # tipo_transaccion por defecto 'A' (AJUSTE) para AD/AC reales (ver
    # aplicar_saldos_menores); reversar_documento pasa el tipo_transaccion
    # real de TCXP_TDOCU para ND/NC ('D'/'C') -- antes quedaba hardcodeado
    # en 'A' para todo, asi que un reverso etiquetado ND se guardaba con
    # tipo_transaccion de "ajuste" en vez de "nota de debito/credito".
    cur.execute(
        "INSERT INTO CXP.TCXP_DOCUMENTO ("
        " no_cia, punto, tipo_docu, no_docu, no_proveedor, tipo_movi,"
        " tipo_transaccion, fecha, status, valor_original, saldo,"
        " detalle, usuario, st_generado_cnt, st_impresion, debito, credito"
        ") VALUES (:1,:2,:3,:4,:5,:6,:7,TO_DATE(:8,'YYYY-MM-DD'),"
        " 'A',:9,0,:10,:11,'N','N',:12,:13)",
        [no_cia, punto, tipo_docu, no_docu, str(no_proveedor), tipo_movi,
         tipo_transaccion, fecha, valor, motivo, usuario, valor, valor])


def _insert_cxp_ajuste_lineas(cur, no_cia, punto, tipo_docu, no_docu,
                               no_proveedor, monto, cuenta_ajuste, centro_costo,
                               cuenta_proveedor, signo_ajuste, signo_proveedor):
    if cuenta_ajuste:
        cur.execute(
            "INSERT INTO CXP.TCXP_DCDOCU ("
            " no_cia, punto, tipo_docu, no_docu, cuenta, centro_costo,"
            " tipo_movi, no_proveedor, monto"
            ") VALUES (:1,:2,:3,:4,:5,:6,:7,:8,:9)",
            [no_cia, punto, tipo_docu, no_docu, cuenta_ajuste,
             centro_costo, signo_ajuste, str(no_proveedor), monto])
    if cuenta_proveedor:
        cur.execute(
            "INSERT INTO CXP.TCXP_DCDOCU ("
            " no_cia, punto, tipo_docu, no_docu, cuenta, centro_costo,"
            " tipo_movi, no_proveedor, monto"
            ") VALUES (:1,:2,:3,:4,:5,'0000000000',:6,:7,:8)",
            [no_cia, punto, tipo_docu, no_docu, cuenta_proveedor,
             signo_proveedor, str(no_proveedor), monto])


def _aplicar_refedocu_cxp(cur, no_cia, punto, tipo_docu, no_docu,
                           no_proveedor, docs_afectados):
    for d in docs_afectados:
        saldo = float(d['saldo'])
        monto = abs(saldo)
        tipo_ref = d['tipo_doc'].strip()
        no_ref = d['no_doc'].strip()
        cur.execute(
            "INSERT INTO CXP.TCXP_REFEDOCU "
            "(no_cia, punto, tipo_docu, no_docu, no_proveedor, tipo_refe, no_refe, monto) "
            "VALUES (:1,:2,:3,:4,:5,:6,:7,:8)",
            [no_cia, punto, tipo_docu, no_docu, str(no_proveedor),
             tipo_ref, no_ref, monto])
        cur.execute(
            "UPDATE CXP.TCXP_DOCUMENTO SET saldo=0, status='C' "
            "WHERE no_cia=:1 AND punto=:2 AND tipo_docu=:3 AND no_docu=:4",
            [no_cia, punto, tipo_ref, no_ref])
